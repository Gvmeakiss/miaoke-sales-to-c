#!/usr/bin/env python3
"""按五月相同口径核对2026年1—2月发货对账明细与OMS月结Y001。"""

from __future__ import annotations

import csv
import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from analyze_may_reconciliation_vs_oms import (
    EXPECTED_HEADERS,
    iter_sheet_rows,
    match_rate,
    normalize_code,
    normalize_shop,
    number,
    parse_row,
)
from oms_transaction_codes import OMS_STANDARD_SETTLEMENT_CODES


ROOT = Path(__file__).resolve().parents[1]
RECON_FILE = ROOT / "input/对账明细（to oms 月结）/发货对账明细2026.01-2026.02.xlsx"
OMS_FILES = {
    "2026-01": ROOT / "input/OMS_2C单据_Excel/oms月结/OMS_月结_26年01月.xlsx",
    "2026-02": ROOT / "input/OMS_2C单据_Excel/oms月结/OMS_月结_26年02月.xlsx",
}
SHOP_MASTER_FILE = ROOT / "input/旺店通内店铺id与店铺名称映射.csv"
OUTPUT_DIR = ROOT / "reconciliation/results"
RESULT_JSON = OUTPUT_DIR / "jan_feb_reconciliation_vs_oms.json"
DETAIL_CSV = OUTPUT_DIR / "jan_feb_reconciliation_vs_oms_customer_item.csv"
TARGET_MONTHS = tuple(OMS_FILES)


def clean_excel_text(value: object) -> str:
    text = str(value or "").strip()
    match = re.fullmatch(r'=\"(.*)\"', text)
    return match.group(1) if match else text


def load_shop_master() -> dict[str, str]:
    result = {}
    with SHOP_MASTER_FILE.open(encoding="utf-8-sig", newline="") as stream:
        for row in csv.DictReader(stream):
            shop_id = clean_excel_text(row.get("店铺编号"))
            shop_name = str(row.get("店铺名称") or "").strip()
            if shop_id and shop_name:
                result[shop_id] = shop_name
    return result


def load_oms():
    aggregate = defaultdict(lambda: {"qty": 0.0, "amount": 0.0, "rows": 0, "customer_name": ""})
    document_nos, business_nos = set(), set()
    by_code = defaultdict(lambda: {"rows": 0, "qty": 0.0, "amount": 0.0})
    for month, path in OMS_FILES.items():
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        indexes = {str(value): index for index, value in enumerate(headers)}
        required = ["商品编码", "数量", "分摊金额", "业务单号", "单据号", "业务类型", "客户编码", "客户名称"]
        missing = [field for field in required if field not in indexes]
        if missing:
            raise ValueError(f"{path.name}缺少字段：{missing}")
        for row in rows:
            business_type = str(row[indexes["业务类型"]] or "").strip()
            if business_type not in OMS_STANDARD_SETTLEMENT_CODES:
                continue
            customer_code = normalize_shop(row[indexes["客户编码"]])
            item_code = normalize_code(row[indexes["商品编码"]])
            if not customer_code or not item_code:
                continue
            key = (month, customer_code, item_code)
            item = aggregate[key]
            item["qty"] += number(row[indexes["数量"]])
            item["amount"] += number(row[indexes["分摊金额"]])
            item["rows"] += 1
            item["customer_name"] = str(row[indexes["客户名称"]] or "").strip()
            document_nos.add(str(row[indexes["单据号"]] or "").strip())
            business_nos.add(str(row[indexes["业务单号"]] or "").strip())
            control = by_code[(month, business_type)]
            control["rows"] += 1
            control["qty"] += number(row[indexes["数量"]])
            control["amount"] += number(row[indexes["分摊金额"]])
        workbook.close()
    return aggregate, document_nos, business_nos, by_code


def month_from_period(start: str, end: str) -> str:
    for value in (start, end):
        match = re.search(r"(20\d{2})[-/.](\d{1,2})", value or "")
        if match:
            return f"{match.group(1)}-{int(match.group(2)):02d}"
    return ""


def load_reconciliation(oms_document_nos: set[str], oms_business_nos: set[str]):
    aggregate = defaultdict(lambda: {"qty": 0.0, "amount": 0.0, "rows": 0})
    status_totals = defaultdict(lambda: {"rows": 0, "qty": 0.0, "amount": 0.0})
    stats = Counter()
    periods = Counter()
    header_checks = []
    with zipfile.ZipFile(RECON_FILE) as archive:
        sheet_names = sorted(
            (name for name in archive.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name)),
            key=lambda name: int(re.search(r"sheet(\d+)", name).group(1)),
        )
        for sheet_name in sheet_names:
            with archive.open(sheet_name) as stream:
                rows = iter_sheet_rows(stream)
                header = parse_row(next(rows))
                mismatches = {
                    column: [EXPECTED_HEADERS[column], header.get(column, "")]
                    for column in EXPECTED_HEADERS
                    if header.get(column, "") != EXPECTED_HEADERS[column]
                }
                header_checks.append({"sheet": sheet_name, "mismatches": mismatches})
                if mismatches:
                    raise ValueError(f"{sheet_name}字段不一致：{mismatches}")
                for row_xml in rows:
                    row = parse_row(row_xml)
                    stats["source_rows"] += 1
                    month = month_from_period(row.get("B", ""), row.get("C", ""))
                    periods[(row.get("B", ""), row.get("C", ""))] += 1
                    if month not in TARGET_MONTHS:
                        stats["out_of_scope_rows"] += 1
                        continue
                    status = row.get("P", "") or "空白"
                    shop_code = normalize_shop(row.get("R"))
                    item_code = normalize_code(row.get("AJ")) or normalize_code(row.get("J"))
                    actual_qty = number(row.get("AF"))
                    receipt_amount = number(row.get("W"))
                    status_item = status_totals[(month, status)]
                    status_item["rows"] += 1
                    status_item["qty"] += actual_qty
                    status_item["amount"] += receipt_amount
                    if any(
                        row.get(column, "") in oms_document_nos or row.get(column, "") in oms_business_nos
                        for column in ("A", "G", "H")
                    ):
                        stats["direct_document_hit_rows"] += 1
                    if status != "对账成功":
                        continue
                    stats["success_rows"] += 1
                    if not shop_code or not item_code:
                        stats["success_blank_key_rows"] += 1
                        continue
                    key = (month, shop_code, item_code)
                    item = aggregate[key]
                    item["qty"] += actual_qty
                    item["amount"] += receipt_amount
                    item["rows"] += 1
    return aggregate, status_totals, stats, periods, header_checks


def totals(dataset, keys):
    return {
        "keys": len(keys),
        "qty": sum(dataset[key]["qty"] for key in keys),
        "amount": sum(dataset[key]["amount"] for key in keys),
    }


def compare(recon, oms, months):
    recon_keys = {key for key in recon if key[0] in months}
    oms_keys = {key for key in oms if key[0] in months}
    common = recon_keys & oms_keys
    only_recon = recon_keys - oms_keys
    only_oms = oms_keys - recon_keys
    left = totals(recon, common)
    right = totals(oms, common)
    return {
        "reconciliation_total": totals(recon, recon_keys),
        "oms_total": totals(oms, oms_keys),
        "common_keys": len(common),
        "common_reconciliation": left,
        "common_oms": right,
        "common_amount_difference": left["amount"] - right["amount"],
        "common_quantity_difference": left["qty"] - right["qty"],
        "common_amount_match_rate": match_rate(left["amount"], right["amount"]),
        "common_quantity_match_rate": match_rate(left["qty"], right["qty"]),
        "exact_amount_keys": sum(abs(recon[key]["amount"] - oms[key]["amount"]) <= 0.01 for key in common),
        "exact_quantity_keys": sum(abs(recon[key]["qty"] - oms[key]["qty"]) <= 0.000001 for key in common),
        "reconciliation_only": totals(recon, only_recon),
        "oms_only": totals(oms, only_oms),
        "sets": {"common": common, "reconciliation_only": only_recon, "oms_only": only_oms},
    }


def public_result(result):
    return {key: value for key, value in result.items() if key != "sets"}


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    shop_master = load_shop_master()
    oms, document_nos, business_nos, oms_by_code = load_oms()
    recon, status_totals, stats, periods, header_checks = load_reconciliation(document_nos, business_nos)
    comparisons = {month: compare(recon, oms, {month}) for month in TARGET_MONTHS}
    comparisons["2026-01至2026-02"] = compare(recon, oms, set(TARGET_MONTHS))
    combined = comparisons["2026-01至2026-02"]

    with DETAIL_CSV.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.writer(stream)
        writer.writerow([
            "匹配分类", "月份", "店铺/客户编码", "旺店通店铺名称", "OMS客户名称", "货品/商品编码",
            "对账清单行数", "对账清单实际数量", "OMS数量", "数量差异",
            "对账清单收款金额", "OMS分摊金额", "金额差异",
        ])
        all_keys = set().union(*combined["sets"].values())
        for key in sorted(all_keys):
            month, shop_code, item_code = key
            left = recon.get(key, {"rows": 0, "qty": 0.0, "amount": 0.0})
            right = oms.get(key, {"qty": 0.0, "amount": 0.0, "customer_name": ""})
            category = "共同键" if key in combined["sets"]["common"] else "仅对账清单" if key in combined["sets"]["reconciliation_only"] else "仅OMS月结"
            writer.writerow([
                category, month, shop_code, shop_master.get(shop_code, ""), right.get("customer_name", ""), item_code,
                left.get("rows", 0), left.get("qty", 0.0), right.get("qty", 0.0), left.get("qty", 0.0) - right.get("qty", 0.0),
                left.get("amount", 0.0), right.get("amount", 0.0), left.get("amount", 0.0) - right.get("amount", 0.0),
            ])

    payload = {
        "scope": {
            "reconciliation_file": str(RECON_FILE.relative_to(ROOT)),
            "oms_files": [str(path.relative_to(ROOT)) for path in OMS_FILES.values()],
            "reconciliation_filter": "对账状态=对账成功；按账期开始日期归属月份",
            "oms_filter": "业务类型=Y001（月结标准结算子集）",
            "key": "月份+店铺编码/客户编码+货品编码/商品编码",
            "amount_fields": "对账清单=收款金额；OMS=分摊金额/share_amount",
            "quantity_fields": "对账清单=实际数量；OMS=数量/item_num",
        },
        "source_checks": {
            "headers": header_checks,
            "periods": [{"start": key[0], "end": key[1], "rows": value} for key, value in periods.most_common()],
            "source_rows": stats["source_rows"],
            "success_rows": stats["success_rows"],
            "success_blank_key_rows": stats["success_blank_key_rows"],
            "out_of_scope_rows": stats["out_of_scope_rows"],
            "direct_document_hit_rows": stats["direct_document_hit_rows"],
            "shop_master_coverage": {
                "reconciliation_shops": len({key[1] for key in recon}),
                "mapped_shops": len({key[1] for key in recon if key[1] in shop_master}),
            },
        },
        "oms_y001_by_month": [
            {"month": month, "business_type": code, **values}
            for (month, code), values in sorted(oms_by_code.items())
        ],
        "status_totals": [
            {"month": month, "status": status, **values}
            for (month, status), values in sorted(status_totals.items())
        ],
        "comparisons": {label: public_result(result) for label, result in comparisons.items()},
    }
    RESULT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(payload, ensure_ascii=False, indent=2))
    print(f"RESULT_JSON={RESULT_JSON}")
    print(f"DETAIL_CSV={DETAIL_CSV}")


if __name__ == "__main__":
    main()
