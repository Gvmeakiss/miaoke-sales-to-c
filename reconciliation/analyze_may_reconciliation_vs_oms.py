#!/usr/bin/env python3
"""五月发货对账清单与OMS月结Y001标准结算专项核对。

对账清单使用账期为2026年5月的《发货对账明细2026.05.xlsx》，
OMS使用《OMS_月结_26年05月.xlsx》中月结事务码Y001标准结算子集。对账清单仅纳入
“对账成功”记录，并按店铺编码+货品编码汇总后与OMS客户编码+商品编码核对。
"""

from __future__ import annotations

import csv
import html
import json
import math
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

from oms_transaction_codes import OMS_STANDARD_SETTLEMENT_CODES


ROOT = Path(__file__).resolve().parents[1]
RECON_FILE = ROOT / "input/对账明细（to oms 月结）/发货对账明细2026.05.xlsx"
OMS_FILE = ROOT / "input/OMS_2C单据_Excel/oms月结/OMS_月结_26年05月.xlsx"
OUTPUT_DIR = ROOT / "reconciliation/results"
RESULT_JSON = OUTPUT_DIR / "may_reconciliation_vs_oms.json"
DETAIL_CSV = OUTPUT_DIR / "may_reconciliation_vs_oms_customer_item.csv"

NS_ROW_END = b"</row>"
NEEDED_COLS = ["A", "B", "C", "G", "H", "J", "K", "P", "R", "S", "T", "W", "X", "Y", "AF", "AG", "AI", "AJ"]
CELL_RE = re.compile(
    rb'<c r="(' + b"|".join(x.encode() for x in sorted(NEEDED_COLS, key=len, reverse=True)) + rb')\d+"[^>]*>(.*?)</c>'
)
VALUE_RE = re.compile(rb"<(?:t|v)(?:\s[^>]*)?>(.*?)</(?:t|v)>")

EXPECTED_HEADERS = {
    "A": "汇总单号", "B": "账期开始日期", "C": "账期结束日期", "G": "系统单号",
    "H": "出库单号", "J": "商品编码", "K": "仓库编码", "P": "对账状态", "R": "店铺编码",
    "S": "发货金额", "T": "调整单金额", "W": "收款金额", "X": "差异金额",
    "Y": "售中退款金额", "AF": "实际数量", "AG": "发货数量", "AI": "退货入库数量",
    "AJ": "货品编码",
}


def decode_value(raw: bytes) -> str:
    match = VALUE_RE.search(raw)
    if not match:
        return ""
    return html.unescape(match.group(1).decode("utf-8", errors="replace")).strip()


def parse_row(row_xml: bytes) -> dict[str, str]:
    return {match.group(1).decode(): decode_value(match.group(2)) for match in CELL_RE.finditer(row_xml)}


def iter_sheet_rows(stream, chunk_size: int = 8 * 1024 * 1024):
    buffer = b""
    while True:
        chunk = stream.read(chunk_size)
        if not chunk:
            break
        buffer += chunk
        while True:
            start = buffer.find(b"<row")
            if start < 0:
                buffer = buffer[-64:]
                break
            end = buffer.find(NS_ROW_END, start)
            if end < 0:
                if start > 0:
                    buffer = buffer[start:]
                break
            end += len(NS_ROW_END)
            yield buffer[start:end]
            buffer = buffer[end:]


def number(value: str | None) -> float:
    try:
        result = float(value or 0)
        return result if math.isfinite(result) else 0.0
    except (TypeError, ValueError):
        return 0.0


def normalize_code(value) -> str:
    text = str(value or "").strip().replace("\u200b", "")
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text


def normalize_shop(value) -> str:
    text = normalize_code(value)
    return text.lstrip("0") or ("0" if text else "")


def match_rate(left: float, right: float) -> float | None:
    denominator = max(abs(left), abs(right))
    return min(abs(left), abs(right)) / denominator if denominator else None


def load_oms():
    workbook = load_workbook(OMS_FILE, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    indexes = {str(value): index for index, value in enumerate(headers)}
    required = ["商品编码", "数量", "销售单位", "分摊金额", "业务单号", "单据号", "业务类型", "客户编码", "客户名称"]
    missing = [field for field in required if field not in indexes]
    if missing:
        raise ValueError(f"OMS月结缺少字段：{missing}")

    aggregate = defaultdict(lambda: {"qty": 0.0, "amount": 0.0, "rows": 0, "units": set(), "customer_name": ""})
    document_nos, business_nos = set(), set()
    total_rows = 0
    for row in rows:
        if str(row[indexes["业务类型"]] or "").strip() not in OMS_STANDARD_SETTLEMENT_CODES:
            continue
        customer_code = normalize_shop(row[indexes["客户编码"]])
        item_code = normalize_code(row[indexes["商品编码"]])
        if not customer_code or not item_code:
            continue
        key = (customer_code, item_code)
        item = aggregate[key]
        item["qty"] += number(row[indexes["数量"]])
        item["amount"] += number(row[indexes["分摊金额"]])
        item["rows"] += 1
        item["units"].add(str(row[indexes["销售单位"]] or "").strip())
        item["customer_name"] = str(row[indexes["客户名称"]] or "").strip()
        document_nos.add(str(row[indexes["单据号"]] or "").strip())
        business_nos.add(str(row[indexes["业务单号"]] or "").strip())
        total_rows += 1
    workbook.close()
    return aggregate, document_nos, business_nos, total_rows


def load_reconciliation(oms_document_nos: set[str], oms_business_nos: set[str]):
    aggregate = defaultdict(lambda: {"qty": 0.0, "ship_qty": 0.0, "amount": 0.0, "ship_amount": 0.0, "rows": 0})
    status_totals = defaultdict(lambda: {"rows": 0, "qty": 0.0, "amount": 0.0})
    stats = Counter()
    periods = Counter()
    header_checks = []

    with zipfile.ZipFile(RECON_FILE) as archive:
        sheet_names = sorted(
            (name for name in archive.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", name)),
            key=lambda name: int(re.search(r"sheet(\d+)", name).group(1)),
        )
        for sheet_name in sheet_names:
            with archive.open(sheet_name) as stream:
                rows = iter_sheet_rows(stream)
                header = parse_row(next(rows))
                mismatches = {column: [EXPECTED_HEADERS[column], header.get(column, "")] for column in EXPECTED_HEADERS if header.get(column, "") != EXPECTED_HEADERS[column]}
                header_checks.append({"sheet": sheet_name, "mismatches": mismatches})
                if mismatches:
                    raise ValueError(f"{sheet_name}字段不一致：{mismatches}")

                for row_xml in rows:
                    row = parse_row(row_xml)
                    stats["source_rows"] += 1
                    status = row.get("P", "") or "空白"
                    shop_code = normalize_shop(row.get("R"))
                    warehouse_code = normalize_code(row.get("K"))
                    item_code = normalize_code(row.get("AJ")) or normalize_code(row.get("J"))
                    if row.get("AJ"):
                        stats["item_from_goods_code"] += 1
                    elif row.get("J"):
                        stats["item_from_product_code_fallback"] += 1
                    else:
                        stats["blank_item_code"] += 1
                    if not shop_code:
                        stats["blank_shop_code"] += 1
                    if warehouse_code:
                        stats[f"warehouse::{warehouse_code}"] += 1

                    actual_qty = number(row.get("AF"))
                    ship_qty = number(row.get("AG"))
                    return_in_qty = number(row.get("AI"))
                    ship_amount = number(row.get("S"))
                    adjustment_amount = number(row.get("T"))
                    receipt_amount = number(row.get("W"))
                    difference_amount = number(row.get("X"))
                    refund_amount = number(row.get("Y"))

                    if abs(actual_qty - (ship_qty - return_in_qty)) > 0.000001:
                        stats["quantity_formula_breaks"] += 1
                    if abs(difference_amount - (ship_amount + adjustment_amount - receipt_amount - refund_amount)) > 0.010001:
                        stats["amount_formula_breaks"] += 1

                    status_item = status_totals[status]
                    status_item["rows"] += 1
                    status_item["qty"] += actual_qty
                    status_item["amount"] += receipt_amount
                    periods[(row.get("B", ""), row.get("C", ""))] += 1

                    if any((row.get(column, "") in oms_document_nos or row.get(column, "") in oms_business_nos) for column in ("A", "G", "H")):
                        stats["direct_document_hit_rows"] += 1

                    if status != "对账成功":
                        continue
                    stats["success_rows"] += 1
                    if not shop_code or not item_code:
                        stats["success_blank_key_rows"] += 1
                        continue
                    key = (shop_code, item_code)
                    item = aggregate[key]
                    item["qty"] += actual_qty
                    item["ship_qty"] += ship_qty
                    item["amount"] += receipt_amount
                    item["ship_amount"] += ship_amount + adjustment_amount - refund_amount
                    item["rows"] += 1

    return aggregate, status_totals, stats, periods, header_checks


def summarize(recon, oms, status_totals, stats, periods, header_checks, oms_rows):
    recon_keys, oms_keys = set(recon), set(oms)
    common_keys = recon_keys & oms_keys
    only_recon = recon_keys - oms_keys
    only_oms = oms_keys - recon_keys

    def totals(dataset, keys):
        return {
            "keys": len(keys),
            "qty": sum(dataset[key]["qty"] for key in keys),
            "amount": sum(dataset[key]["amount"] for key in keys),
        }

    recon_total = totals(recon, recon_keys)
    oms_total = totals(oms, oms_keys)
    recon_common = totals(recon, common_keys)
    oms_common = totals(oms, common_keys)
    recon_only = totals(recon, only_recon)
    oms_only = totals(oms, only_oms)

    exact_amount = sum(1 for key in common_keys if abs(recon[key]["amount"] - oms[key]["amount"]) <= 0.01)
    exact_qty = sum(1 for key in common_keys if abs(recon[key]["qty"] - oms[key]["qty"]) <= 0.000001)
    exact_both = sum(1 for key in common_keys if abs(recon[key]["amount"] - oms[key]["amount"]) <= 0.01 and abs(recon[key]["qty"] - oms[key]["qty"]) <= 0.000001)
    exact_ship_amount = sum(1 for key in common_keys if abs(recon[key]["ship_amount"] - oms[key]["amount"]) <= 0.01)
    exact_ship_qty = sum(1 for key in common_keys if abs(recon[key]["ship_qty"] - oms[key]["qty"]) <= 0.000001)
    common_ship_amount = sum(recon[key]["ship_amount"] for key in common_keys)
    common_ship_qty = sum(recon[key]["ship_qty"] for key in common_keys)

    recon_shop = defaultdict(lambda: {"amount": 0.0, "qty": 0.0})
    oms_shop = defaultdict(lambda: {"amount": 0.0, "qty": 0.0})
    for (shop, _), values in recon.items():
        recon_shop[shop]["amount"] += values["amount"]
        recon_shop[shop]["qty"] += values["qty"]
    for (shop, _), values in oms.items():
        oms_shop[shop]["amount"] += values["amount"]
        oms_shop[shop]["qty"] += values["qty"]
    common_shops = set(recon_shop) & set(oms_shop)
    recon_shop_amount = sum(recon_shop[key]["amount"] for key in common_shops)
    oms_shop_amount = sum(oms_shop[key]["amount"] for key in common_shops)
    recon_shop_qty = sum(recon_shop[key]["qty"] for key in common_shops)
    oms_shop_qty = sum(oms_shop[key]["qty"] for key in common_shops)

    recon_item = defaultdict(lambda: {"amount": 0.0, "qty": 0.0})
    oms_item = defaultdict(lambda: {"amount": 0.0, "qty": 0.0})
    for (_, item_code), values in recon.items():
        recon_item[item_code]["amount"] += values["amount"]
        recon_item[item_code]["qty"] += values["qty"]
    for (_, item_code), values in oms.items():
        oms_item[item_code]["amount"] += values["amount"]
        oms_item[item_code]["qty"] += values["qty"]
    common_items = set(recon_item) & set(oms_item)
    recon_item_amount = sum(recon_item[key]["amount"] for key in common_items)
    oms_item_amount = sum(oms_item[key]["amount"] for key in common_items)
    recon_item_qty = sum(recon_item[key]["qty"] for key in common_items)
    oms_item_qty = sum(oms_item[key]["qty"] for key in common_items)

    result = {
        "scope": {
            "reconciliation_file": str(RECON_FILE.relative_to(ROOT)),
            "oms_file": str(OMS_FILE.relative_to(ROOT)),
            "reconciliation_filter": "对账状态=对账成功；按账期文件全量；成功记录优先",
            "oms_filter": "业务类型=Y001（事务码表定义的月结标准结算子集）；OMS五月月结导出",
            "reconciliation_key": "店铺编码+货品编码",
            "oms_key": "客户编码+商品编码",
            "amount_fields": "对账清单=收款金额；OMS=分摊金额/share_amount",
            "quantity_fields": "对账清单=实际数量；OMS=数量/item_num",
        },
        "source_checks": {
            "headers": header_checks,
            "periods": [{"start": key[0], "end": key[1], "rows": value} for key, value in periods.most_common()],
            "source_rows": stats["source_rows"],
            "success_rows": stats["success_rows"],
            "success_blank_key_rows": stats["success_blank_key_rows"],
            "quantity_formula_breaks": stats["quantity_formula_breaks"],
            "amount_formula_breaks": stats["amount_formula_breaks"],
            "direct_document_hit_rows": stats["direct_document_hit_rows"],
            "oms_y001_rows": oms_rows,
            "oms_multi_unit_customer_item_keys": sum(1 for value in oms.values() if len(value["units"]) > 1),
            "reconciliation_warehouses": [
                {"warehouse_code": key.split("::", 1)[1], "rows": value}
                for key, value in sorted(stats.items()) if key.startswith("warehouse::")
            ],
        },
        "status_totals": [
            {"status": status, "rows": values["rows"], "qty": values["qty"], "amount": values["amount"]}
            for status, values in sorted(status_totals.items())
        ],
        "total_comparison": {
            "reconciliation": recon_total,
            "oms": oms_total,
            "amount_difference_reconciliation_minus_oms": recon_total["amount"] - oms_total["amount"],
            "quantity_difference_reconciliation_minus_oms": recon_total["qty"] - oms_total["qty"],
            "amount_match_rate": match_rate(recon_total["amount"], oms_total["amount"]),
            "quantity_match_rate": match_rate(recon_total["qty"], oms_total["qty"]),
        },
        "common_customer_item": {
            "common_keys": len(common_keys),
            "reconciliation": recon_common,
            "oms": oms_common,
            "amount_difference_reconciliation_minus_oms": recon_common["amount"] - oms_common["amount"],
            "quantity_difference_reconciliation_minus_oms": recon_common["qty"] - oms_common["qty"],
            "amount_match_rate": match_rate(recon_common["amount"], oms_common["amount"]),
            "quantity_match_rate": match_rate(recon_common["qty"], oms_common["qty"]),
            "reconciliation_key_coverage": len(common_keys) / len(recon_keys) if recon_keys else None,
            "oms_key_coverage": len(common_keys) / len(oms_keys) if oms_keys else None,
            "exact_amount_keys": exact_amount,
            "exact_quantity_keys": exact_qty,
            "exact_amount_and_quantity_keys": exact_both,
            "alternative_fields": {
                "发货调整退款净额_vs_OMS分摊金额": {
                    "reconciliation_amount": common_ship_amount,
                    "oms_amount": oms_common["amount"],
                    "amount_match_rate": match_rate(common_ship_amount, oms_common["amount"]),
                    "exact_amount_keys": exact_ship_amount,
                },
                "发货数量_vs_OMS数量": {
                    "reconciliation_qty": common_ship_qty,
                    "oms_qty": oms_common["qty"],
                    "quantity_match_rate": match_rate(common_ship_qty, oms_common["qty"]),
                    "exact_quantity_keys": exact_ship_qty,
                },
            },
        },
        "one_sided": {
            "reconciliation_only": recon_only,
            "oms_only": oms_only,
        },
        "common_shop": {
            "common_shops": len(common_shops),
            "reconciliation_amount": recon_shop_amount,
            "oms_amount": oms_shop_amount,
            "amount_match_rate": match_rate(recon_shop_amount, oms_shop_amount),
            "reconciliation_qty": recon_shop_qty,
            "oms_qty": oms_shop_qty,
            "quantity_match_rate": match_rate(recon_shop_qty, oms_shop_qty),
        },
        "common_item": {
            "common_items": len(common_items),
            "reconciliation_amount": recon_item_amount,
            "oms_amount": oms_item_amount,
            "amount_match_rate": match_rate(recon_item_amount, oms_item_amount),
            "reconciliation_qty": recon_item_qty,
            "oms_qty": oms_item_qty,
            "quantity_match_rate": match_rate(recon_item_qty, oms_item_qty),
        },
    }
    return result, common_keys, only_recon, only_oms


def write_detail(recon, oms, common_keys, only_recon, only_oms):
    with DETAIL_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow([
            "匹配分类", "店铺/客户编码", "货品/商品编码", "OMS客户名称", "对账清单行数",
            "对账清单实际数量", "OMS数量", "数量差异", "对账清单收款金额", "OMS分摊金额", "金额差异",
        ])
        for key in sorted(common_keys | only_recon | only_oms):
            left = recon.get(key, {"rows": 0, "qty": 0.0, "amount": 0.0})
            right = oms.get(key, {"qty": 0.0, "amount": 0.0, "customer_name": ""})
            category = "共同键" if key in common_keys else "仅对账清单" if key in only_recon else "仅OMS月结"
            writer.writerow([
                category, key[0], key[1], right.get("customer_name", ""), left.get("rows", 0),
                left.get("qty", 0.0), right.get("qty", 0.0), left.get("qty", 0.0) - right.get("qty", 0.0),
                left.get("amount", 0.0), right.get("amount", 0.0), left.get("amount", 0.0) - right.get("amount", 0.0),
            ])


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    oms, document_nos, business_nos, oms_rows = load_oms()
    recon, status_totals, stats, periods, header_checks = load_reconciliation(document_nos, business_nos)
    result, common_keys, only_recon, only_oms = summarize(recon, oms, status_totals, stats, periods, header_checks, oms_rows)
    write_detail(recon, oms, common_keys, only_recon, only_oms)
    RESULT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    print(f"RESULT_JSON={RESULT_JSON}")
    print(f"DETAIL_CSV={DETAIL_CSV}")


if __name__ == "__main__":
    main()
