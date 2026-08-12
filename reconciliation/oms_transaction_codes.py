"""OMS日结/月结事务码的统一识别规则。

权威来源：input/日结月结事务码.xlsx（Sheet1!A1:B15）。
`cycle_type`是OMS源文件的“结算类型”字段，不用于判断日结/月结。
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE_FILE = ROOT / "input" / "日结月结事务码.xlsx"

# 当权威工作簿暂不可用时使用的审计留痕回退值；正常运行会读取工作簿并校验。
FALLBACK_DAY_CODES = (
    "Y005", "Y006", "Y002", "Z001", "Z002", "Z003", "Z004", "Z005", "Z006",
    "Y051", *tuple(f"Z{number:03d}" for number in range(11, 52)), "Y052", "Y021",
)
FALLBACK_MONTH_CODES = ("Z008", "Y001", "Y003", "Y004", "Z007", "Y011")

# 专项核对范围不是日/月结分类本身：标准发票（2C）共同键仅落在Y001；
# 旺店通前端正常销售发货专项核对仅使用Y005。其他事务码仍保留在日/月结全量中。
OMS_STANDARD_SETTLEMENT_CODES = ("Y001",)
OMS_WDT_SALES_DAY_CODES = ("Y005",)


def expand_code(value: object) -> tuple[str, ...]:
    text = str(value or "").strip().upper().replace("—", "-").replace("–", "-")
    match = re.fullmatch(r"([A-Z])(\d+)-([A-Z])(\d+)", text)
    if not match:
        return (text,) if text else ()
    left_prefix, left_number, right_prefix, right_number = match.groups()
    if left_prefix != right_prefix:
        raise ValueError(f"无法展开跨前缀事务码范围：{text}")
    width = max(len(left_number), len(right_number))
    start, end = int(left_number), int(right_number)
    if end < start:
        raise ValueError(f"事务码范围结束值小于起始值：{text}")
    return tuple(f"{left_prefix}{number:0{width}d}" for number in range(start, end + 1))


def load_code_map(source: Path = SOURCE_FILE) -> dict[str, tuple[str, ...]]:
    if not source.exists():
        return {"日结": FALLBACK_DAY_CODES, "月结": FALLBACK_MONTH_CODES}
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        rows = sheet.iter_rows(min_row=2, values_only=True)
        grouped: dict[str, list[str]] = {"日结": [], "月结": []}
        for raw_code, raw_type, *_ in rows:
            cycle = str(raw_type or "").strip()
            if cycle not in grouped:
                continue
            grouped[cycle].extend(expand_code(raw_code))
        return {cycle: tuple(dict.fromkeys(codes)) for cycle, codes in grouped.items()}
    finally:
        workbook.close()


OMS_CODE_MAP = load_code_map()
OMS_DAY_CODES = OMS_CODE_MAP["日结"]
OMS_MONTH_CODES = OMS_CODE_MAP["月结"]


def classify_transaction_code(code: object) -> str:
    normalized = str(code or "").strip().upper()
    if normalized in OMS_DAY_CODES:
        return "日结"
    if normalized in OMS_MONTH_CODES:
        return "月结"
    return "未分类"


def sql_list(codes: Iterable[str]) -> str:
    return ",".join("'" + code.replace("'", "''") + "'" for code in codes)

