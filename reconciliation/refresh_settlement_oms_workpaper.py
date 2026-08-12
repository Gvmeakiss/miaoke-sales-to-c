#!/usr/bin/env python3
"""自动扫描发货对账明细和OMS月结，刷新底稿接口。"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import math
import re
import zipfile
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from oms_transaction_codes import OMS_STANDARD_SETTLEMENT_CODES

ROOT = Path(__file__).resolve().parents[1]
RECON_DIR = ROOT / "input/对账明细（to oms 月结）"
OMS_DIR = ROOT / "input/OMS_2C单据_Excel/oms月结"
SHOP_MASTER = ROOT / "input/旺店通内店铺id与店铺名称映射.csv"
OUTPUT = ROOT / "reconciliation/results/settlement_oms_workpaper.json"
MONTHS = [f"2026-{m:02d}" for m in range(1, 7)]
AUDIT_END = "2026-06-30"

RECON_RE = re.compile(r"^发货对账明细(?P<y>20\d{2})\.(?P<m>\d{2})(?:-(?P<ey>20\d{2})\.(?P<em>\d{2})|-(?P<part>\d+))?\.xlsx$")
RECON_CROSS_MONTH_RE = re.compile(r"^发货对账明细(?P<y>20\d{2})-(?P<m>\d{2})[ _](?P<em>\d{2})\.xlsx$")
OMS_RE = re.compile(r"^OMS_月结_(?P<y>\d{2})年(?P<m>\d{2})月\.xlsx$")
COLS = ["A","B","C","D","E","F","G","H","I","J","K","P","Q","R","W","Y","AF","AI","AJ","AR"]
SIG_COLS = ["A","B","C","D","E","F","G","H","I","J","K","P","R","W","Y","AF","AI","AJ"]
CELL_RE = re.compile(rb'<c r="(' + b"|".join(x.encode() for x in sorted(COLS, key=len, reverse=True)) + rb')\d+"[^>]*>(.*?)</c>')
VALUE_RE = re.compile(rb"<(?:t|v)(?:\s[^>]*)?>(.*?)</(?:t|v)>")
EXPECTED = {
    "A":"汇总单号","B":"账期开始日期","C":"账期结束日期","D":"业务时间","E":"平台订单号",
    "F":"平台子订单号","G":"系统单号","H":"出库单号","I":"退单号","J":"商品编码","K":"仓库编码",
    "P":"对账状态","Q":"店铺名称","R":"店铺编码","W":"收款金额","Y":"售中退款金额",
    "AF":"实际数量","AI":"退货入库数量","AJ":"货品编码","AR":"对账时间",
}


def num(value):
    try:
        value = float(value or 0)
        return value if math.isfinite(value) else 0.0
    except (TypeError, ValueError):
        return 0.0


def norm(value):
    text = str(value or "").strip().replace("\u200b", "")
    return text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text


def norm_shop(value):
    text = norm(value)
    return text.lstrip("0") or ("0" if text else "")


def rate(left, right):
    denominator = max(abs(left), abs(right))
    return min(abs(left), abs(right)) / denominator if denominator else None


def month_of(value):
    text = str(value or "").strip()
    match = re.search(r"(20\d{2})[-/.年](\d{1,2})", text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"
    try:
        serial = float(text)
        if 40000 <= serial <= 60000:
            return (datetime(1899, 12, 30) + timedelta(days=serial)).strftime("%Y-%m")
    except ValueError:
        pass
    return ""


def decode(raw):
    match = VALUE_RE.search(raw)
    return html.unescape(match.group(1).decode("utf-8", "replace")).strip() if match else ""


def parse_row(xml):
    return {m.group(1).decode(): decode(m.group(2)) for m in CELL_RE.finditer(xml)}


def rows_from_stream(stream, chunk_size=8 * 1024 * 1024):
    buffer = b""
    while True:
        chunk = stream.read(chunk_size)
        if not chunk:
            break
        buffer += chunk
        while True:
            start = buffer.find(b"<row")
            if start < 0:
                buffer = buffer[-64:]
                break
            end = buffer.find(b"</row>", start)
            if end < 0:
                buffer = buffer[start:]
                break
            end += 6
            yield buffer[start:end]
            buffer = buffer[end:]


def discover():
    recon = []
    for path in RECON_DIR.glob("*.xlsx") if RECON_DIR.exists() else []:
        match = RECON_RE.fullmatch(path.name)
        cross_month_match = RECON_CROSS_MONTH_RE.fullmatch(path.name)
        if path.name.startswith("~$") or (not match and not cross_month_match):
            continue
        if cross_month_match:
            start = f"{cross_month_match.group('y')}-{int(cross_month_match.group('m')):02d}"
            end = f"{cross_month_match.group('y')}-{int(cross_month_match.group('em')):02d}"
            # 文件名仅用于发现文件；正式归属统一使用原始字段“账期结束日期”。
            recon.append({"path":path, "start":start, "end":end, "range":True, "part":0, "month_basis":"period_end"})
        else:
            start = f"{match.group('y')}-{int(match.group('m')):02d}"
            end = f"{match.group('ey')}-{int(match.group('em')):02d}" if match.group("ey") else start
            # 单月文件、跨月文件及补充分片均可能包含跨期结算记录；不能以文件名月份代替账期字段。
            recon.append({"path":path, "start":start, "end":end, "range":bool(match.group("ey")), "part":int(match.group("part") or 0), "month_basis":"period_end"})
    recon.sort(key=lambda x: (x["start"], x["part"], x["path"].name))
    oms = {}
    for path in OMS_DIR.glob("*.xlsx") if OMS_DIR.exists() else []:
        match = OMS_RE.fullmatch(path.name)
        if match:
            month = f"20{match.group('y')}-{int(match.group('m')):02d}"
            if month in MONTHS:
                oms[month] = path
    return recon, oms


def file_digest(path):
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def signature(row):
    raw = "\x1f".join(row.get(c, "") for c in SIG_COLS).encode("utf-8", "replace")
    return int.from_bytes(hashlib.blake2b(raw, digest_size=8).digest(), "big")


def bucket():
    return {"rows":0,"amount":0.0,"qty":0.0,"files":set(),"periods":set(),"business_months":set(),"recon_months":set(),"shop_name":""}


def load_recon(files):
    data = defaultdict(bucket)
    prior_signatures = set()
    active_file_group = None
    file_hashes = {}
    stats = []
    for info in files:
        path = info["path"]
        # 业务行去重只需覆盖同月分片/补充文件；切换月份后释放签名，避免超大文件累计占用内存。
        if info["start"] != active_file_group:
            prior_signatures.clear()
            active_file_group = info["start"]
        digest = file_digest(path)
        if digest in file_hashes:
            stats.append({"file":path.name,"status":"文件内容重复，未纳入","duplicate_of":file_hashes[digest],"source_rows":0,"success_rows":0,"duplicate_rows_removed":0})
            continue
        file_hashes[digest] = path.name
        current_signatures = set()
        stat = {"file":path.name,"status":"已纳入","month_basis":info.get("month_basis"),"source_rows":0,"success_rows":0,"duplicate_rows_removed":0,"outside_audit_period_rows":0,"outside_audit_period_success_rows":0,"outside_audit_period_success_amount":0.0,"outside_audit_period_success_quantity":0.0}
        with zipfile.ZipFile(path) as archive:
            sheets = sorted((n for n in archive.namelist() if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", n)), key=lambda n:int(re.search(r"sheet(\d+)", n).group(1)))
            for sheet_name in sheets:
                with archive.open(sheet_name) as stream:
                    rows = rows_from_stream(stream)
                    try:
                        header = parse_row(next(rows))
                    except StopIteration:
                        continue
                    mismatch = {c:[name,header.get(c,"")] for c,name in EXPECTED.items() if header.get(c,"") != name}
                    if mismatch:
                        raise ValueError(f"{path.name}/{sheet_name}字段不一致：{mismatch}")
                    for xml in rows:
                        stat["source_rows"] += 1
                        row = parse_row(xml)
                        month = month_of(row.get("C"))
                        if month not in MONTHS:
                            stat["outside_audit_period_rows"] += 1
                            if row.get("P") == "对账成功":
                                stat["outside_audit_period_success_rows"] += 1
                                stat["outside_audit_period_success_amount"] += num(row.get("W"))
                                stat["outside_audit_period_success_quantity"] += num(row.get("AF"))
                            continue
                        if row.get("P") != "对账成功":
                            continue
                        stat["success_rows"] += 1
                        sig = signature(row)
                        if sig in prior_signatures:
                            stat["duplicate_rows_removed"] += 1
                            continue
                        current_signatures.add(sig)
                        shop = norm_shop(row.get("R"))
                        item = norm(row.get("AJ")) or norm(row.get("J"))
                        if not shop or not item:
                            continue
                        value = data[(month, shop, item)]
                        value["rows"] += 1
                        value["amount"] += num(row.get("W"))
                        value["qty"] += num(row.get("AF"))
                        value["files"].add(path.name)
                        value["periods"].add(f"{row.get('B','')}至{row.get('C','')}")
                        business_month, recon_month = month_of(row.get("D")), month_of(row.get("AR"))
                        if business_month: value["business_months"].add(business_month)
                        if recon_month: value["recon_months"].add(recon_month)
                        if row.get("Q") and not value["shop_name"]: value["shop_name"] = row["Q"]
        prior_signatures.update(current_signatures)
        stats.append(stat)
    return data, stats


def load_oms(files):
    data = defaultdict(lambda:{"rows":0,"amount":0.0,"qty":0.0,"customer_name":""})
    stats = []
    for month, path in sorted(files.items()):
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        headers = next(rows)
        index = {str(v):i for i,v in enumerate(headers)}
        missing = [x for x in ["商品编码","数量","分摊金额","业务类型","客户编码","客户名称"] if x not in index]
        if missing:
            workbook.close()
            raise ValueError(f"{path.name}缺少字段：{missing}")
        count = 0
        for row in rows:
            if str(row[index["业务类型"]] or "").strip() not in OMS_STANDARD_SETTLEMENT_CODES:
                continue
            shop, item = norm_shop(row[index["客户编码"]]), norm(row[index["商品编码"]])
            if not shop or not item: continue
            value = data[(month, shop, item)]
            value["rows"] += 1
            value["amount"] += num(row[index["分摊金额"]])
            value["qty"] += num(row[index["数量"]])
            if not value["customer_name"]: value["customer_name"] = str(row[index["客户名称"]] or "").strip()
            count += 1
        workbook.close()
        stats.append({"month":month,"file":path.name,"y001_rows":count})
    return data, stats


def shop_master():
    result = {}
    if SHOP_MASTER.exists():
        with SHOP_MASTER.open(encoding="utf-8-sig", newline="") as stream:
            for row in csv.DictReader(stream):
                code, name = norm_shop(row.get("店铺编号")), str(row.get("店铺名称") or "").strip()
                if code and name: result[code] = name
    return result


def cross_period(month, value):
    return any(x and x != month for x in value["business_months"] | value["recon_months"])


def total(data, keys):
    return {"keys":len(keys),"rows":sum(data[k]["rows"] for k in keys),"amount":sum(data[k]["amount"] for k in keys),"qty":sum(data[k]["qty"] for k in keys)}


def build_payload(recon, oms, recon_stats, oms_stats, recon_files, oms_files):
    names = shop_master()
    summary, details = [], []
    for month in MONTHS:
        left_keys = {k for k in recon if k[0] == month}
        right_keys = {k for k in oms if k[0] == month}
        if not left_keys or not right_keys:
            summary.append({"month":month,"status":"待获取发货对账明细" if not left_keys else "待获取OMS月结","match_dimension":"月份+店铺/客户编码+货品/商品编码","source_files":"；".join(sorted({f for k in left_keys for f in recon[k]["files"]}))})
            continue
        common, only_left, only_right = left_keys & right_keys, left_keys - right_keys, right_keys - left_keys
        lt, rt, lc, rc = total(recon,left_keys), total(oms,right_keys), total(recon,common), total(oms,common)
        lo, ro = total(recon,only_left), total(oms,only_right)
        # 客户已确认共同键金额差异来源于跨期结算。原始金额保留不变，仅对已具备跨期证据
        # （业务月份或实际对账月份与账期归属月不同）的共同键建立可追溯桥接调整。
        cross_period_adjustment = sum(-(recon[k]["amount"] - oms[k]["amount"]) for k in common if cross_period(month,recon[k]))
        adjusted_common_amount = lc["amount"] + cross_period_adjustment
        adjusted_common_difference = adjusted_common_amount - rc["amount"]
        files = sorted({f for k in left_keys for f in recon[k]["files"]})
        periods = sorted({p for k in left_keys for p in recon[k]["periods"]})
        cross_files = sorted({f for k in left_keys if cross_period(month,recon[k]) for f in recon[k]["files"]})
        duplicate_rows = sum(s["duplicate_rows_removed"] for s in recon_stats if s["file"] in files)
        summary.append({
            "month":month,"status":"已完成核对（含跨期补充）" if cross_files else "已完成核对","match_dimension":"月份+店铺/客户编码+货品/商品编码",
            "reconciliation_total_amount":lt["amount"],"oms_total_amount":rt["amount"],"reconciliation_common_amount":lc["amount"],"oms_common_amount":rc["amount"],
            "common_amount_difference":lc["amount"]-rc["amount"],"common_amount_match_rate":rate(lc["amount"],rc["amount"]),"oms_amount_coverage":abs(rc["amount"])/max(abs(rt["amount"]),1),
            "cross_period_adjustment_amount":cross_period_adjustment,"adjusted_reconciliation_common_amount":adjusted_common_amount,
            "adjusted_common_amount_difference":adjusted_common_difference,"adjusted_common_amount_match_rate":rate(adjusted_common_amount,rc["amount"]),
            "total_amount_difference":lt["amount"]-rt["amount"],"total_amount_match_rate":rate(lt["amount"],rt["amount"]),
            "reconciliation_total_quantity":lt["qty"],"oms_total_quantity":rt["qty"],"reconciliation_common_quantity":lc["qty"],"oms_common_quantity":rc["qty"],
            "common_quantity_difference":lc["qty"]-rc["qty"],"common_quantity_match_rate":rate(lc["qty"],rc["qty"]),"common_keys":len(common),
            "exact_amount_keys":sum(abs(recon[k]["amount"]-oms[k]["amount"])<=.01 for k in common),"exact_quantity_keys":sum(abs(recon[k]["qty"]-oms[k]["qty"])<=1e-6 for k in common),
            "reconciliation_only_amount":lo["amount"],"oms_only_amount":ro["amount"],"source_files":"；".join(files),"settlement_periods":"；".join(periods),
            "cross_period_files":"；".join(cross_files),"duplicate_rows_removed":duplicate_rows,
        })
        for key in sorted(left_keys | right_keys):
            _, shop, item = key
            left = recon.get(key, bucket())
            right = oms.get(key, {"rows":0,"amount":0.0,"qty":0.0,"customer_name":""})
            la,ra,lq,rq = left["amount"],right["amount"],left["qty"],right["qty"]
            is_cross_period = bool(left["rows"] and cross_period(month,left))
            adjustment = -(la-ra) if key in common and is_cross_period else 0.0
            adjusted_amount = la + adjustment
            if key in common:
                category = "共同键"
                result = "金额数量一致" if abs(la-ra)<=.01 and abs(lq-rq)<=1e-6 else ("金额一致数量差异" if abs(la-ra)<=.01 else "金额差异")
            elif key in only_left: category,result = "仅发货对账","仅发货对账"
            else: category,result = "仅OMS月结","仅OMS月结"
            details.append({
                "month":month,"match_category":category,"result":result,"shop_customer_code":shop,"wdt_shop_name":names.get(shop) or left.get("shop_name", ""),
                "oms_customer_name":right.get("customer_name", ""),"material_code":item,"reconciliation_rows":left["rows"],"reconciliation_amount":la,"oms_amount":ra,
                "amount_difference":la-ra,"amount_match_rate":rate(la,ra),"reconciliation_quantity":lq,"oms_quantity":rq,"quantity_difference":lq-rq,"quantity_match_rate":rate(lq,rq),
                "cross_period_adjustment_amount":adjustment,"adjusted_reconciliation_amount":adjusted_amount,
                "adjusted_amount_difference":adjusted_amount-ra,"adjusted_amount_match_rate":rate(adjusted_amount,ra),
                "source_files":"；".join(sorted(left["files"])),"settlement_periods":"；".join(sorted(left["periods"])),"business_months":"；".join(sorted(left["business_months"])),
                "cross_period_flag":"是" if is_cross_period else "否",
                "adjustment_basis":"客户确认跨期结算；共同键差异作跨期桥接调整，不修改原始金额" if adjustment else "无跨期金额调整",
            })
    available = [x for x in summary if x["status"].startswith("已完成核对")]
    fields = ["reconciliation_total_amount","oms_total_amount","reconciliation_common_amount","oms_common_amount","cross_period_adjustment_amount","adjusted_reconciliation_common_amount","reconciliation_total_quantity","oms_total_quantity","reconciliation_common_quantity","oms_common_quantity","common_keys","exact_amount_keys","exact_quantity_keys","reconciliation_only_amount","oms_only_amount"]
    overall = {f:sum(num(x.get(f)) for x in available) for f in fields}
    overall.update({
        "available_months":[x["month"] for x in available],"pending_months":[x["month"] for x in summary if not x["status"].startswith("已完成核对")],
        "common_amount_difference":overall["reconciliation_common_amount"]-overall["oms_common_amount"],"common_amount_match_rate":rate(overall["reconciliation_common_amount"],overall["oms_common_amount"]),
        "adjusted_common_amount_difference":overall["adjusted_reconciliation_common_amount"]-overall["oms_common_amount"],
        "adjusted_common_amount_match_rate":rate(overall["adjusted_reconciliation_common_amount"],overall["oms_common_amount"]),
        "oms_amount_coverage":abs(overall["oms_common_amount"])/max(abs(overall["oms_total_amount"]),1),"common_quantity_difference":overall["reconciliation_common_quantity"]-overall["oms_common_quantity"],
        "common_quantity_match_rate":rate(overall["reconciliation_common_quantity"],overall["oms_common_quantity"]),"cross_period_files":sorted({x["cross_period_files"] for x in available if x.get("cross_period_files")}),
        "duplicate_rows_removed":sum(int(x.get("duplicate_rows_removed") or 0) for x in available),
    })
    headers = list(details[0]) if details else []
    return {
        "definitions":{"audit_period":"2026-01-01至2026-06-30","reconciliation_filter":"对账状态=对账成功","oms_filter":"业务类型=Y001（月结标准结算子集）","match_dimension":"月份+店铺编码/OMS客户编码+货品编码/OMS商品编码","amount_fields":"发货对账明细.收款金额 vs OMS月结.分摊金额（share_amount）","quantity_fields":"发货对账明细.实际数量 vs OMS月结.数量（item_num）","month_policy":"所有发货对账记录统一使用原始字段“账期结束日期”归属结算月，不以文件名或账期开始月替代；仅纳入账期结束日期位于2026-01-01至2026-06-30的记录；允许6月末D-1/D-2记录在7月初完成对账，账期结束在7月的记录不纳入","cross_period_adjustment_policy":"客户确认差异由跨期结算导致；对业务月份或实际对账月份与账期归属月不同的共同键，将原共同键金额差异作为跨期桥接调整。原始收款金额、原始差异及原账期均保留，调整不覆盖源数据；单边记录不作无依据调整","dedup_policy":"相同文件SHA-256仅纳入一次；跨文件相同业务行仅保留首次出现记录；同一文件内重复行不擅自删除","missing_period_policy":"缺少任一侧资料的月份仅标记待获取，不以0参与合计或匹配率"},
        "source_inventory":{"reconciliation_files":recon_stats,"oms_files":oms_stats,"discovered_reconciliation_files":[x["path"].name for x in recon_files],"discovered_oms_months":sorted(oms_files)},
        "summary_rows":summary,"available_total":overall,"detail_headers":headers,"detail_rows":[[x.get(h) for h in headers] for x in details],
    }


def inventory(recon_files, oms_files):
    by_month = defaultdict(list)
    for info in recon_files:
        if info["range"]:
            y,m = map(int,info["start"].split("-")); ey,em = map(int,info["end"].split("-"))
            while (y,m) <= (ey,em):
                month = f"{y:04d}-{m:02d}"
                if month in MONTHS: by_month[month].append(info["path"].name)
                m += 1
                if m == 13: y,m = y+1,1
        else: by_month[info["start"]].append(info["path"].name)
    return {"reconciliation_directory":str(RECON_DIR),"oms_directory":str(OMS_DIR),"months":[{"month":m,"reconciliation_files":by_month.get(m,[]),"oms_file":oms_files[m].name if m in oms_files else None,"ready":bool(by_month.get(m)) and m in oms_files} for m in MONTHS]}


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--inventory-only", action="store_true", help="仅检查文件发现和期间覆盖，不读取大文件")
    parser.add_argument("--output", type=Path, default=OUTPUT)
    args = parser.parse_args()
    recon_files, oms_files = discover()
    if args.inventory_only:
        print(json.dumps(inventory(recon_files,oms_files),ensure_ascii=False,indent=2)); return
    recon,recon_stats = load_recon(recon_files)
    oms,oms_stats = load_oms(oms_files)
    payload = build_payload(recon,oms,recon_stats,oms_stats,recon_files,oms_files)
    args.output.parent.mkdir(parents=True,exist_ok=True)
    args.output.write_text(json.dumps(payload,ensure_ascii=False,indent=2),encoding="utf-8")
    print(json.dumps({"output":str(args.output),"available_months":payload["available_total"]["available_months"],"pending_months":payload["available_total"]["pending_months"],"detail_rows":len(payload["detail_rows"]),"cross_period_files":payload["available_total"]["cross_period_files"],"duplicate_rows_removed":payload["available_total"]["duplicate_rows_removed"]},ensure_ascii=False,indent=2))


if __name__ == "__main__":
    main()
