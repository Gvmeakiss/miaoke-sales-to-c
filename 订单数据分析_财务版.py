"""
旺店通订单财务分析（2026 平铺分卷导出版）

适用输入：
    input/旺店通订单清单/26年1月-1.xlsx
    input/旺店通订单清单/26年1月-2.xlsx
    input/旺店通订单清单/已归档订单明细*.xlsx

核心原则：
1. 先将商品明细归并为订单级事实，避免「应收金额/订单优惠/订单邮费」因多商品行被重复累加。
2. 分析销售额 = 分摊后总价 + 分摊邮费；同时保留系统应收与已付做对账。
3. 实际退款金额不在当前字段中，不擅自冲减收入；只披露「退款标记订单数/金额」。
4. 成本优先取行级「货品当前成本」，其次「成本价×数量」；无行级成本时才回退订单「预估货品成本」。
5. 默认用发货时间归期，这是经营分析代理口径，不等同会计准则下的法定收入确认日。

用法：
    python3 订单数据分析_财务版.py
    python3 订单数据分析_财务版.py --sample-rows 30000
    python3 订单数据分析_财务版.py --period-basis payment
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import traceback
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT = SCRIPT_DIR / "input" / "旺店通订单清单"
DEFAULT_OUTPUT = SCRIPT_DIR / "output"
DEFAULT_CACHE = SCRIPT_DIR / "PKL_财务版"
CACHE_VERSION = "finance-v1.0"

PERIOD_COLUMNS = {
    "transaction": "交易时间",
    "payment": "付款时间",
    "shipment": "发货时间",
    "finish": "结束时间",
}

COMMERCIAL_ORDER_TYPES = {"网店销售", "分销订单", "线下订单"}
NON_COMMERCIAL_ORDER_TYPES = {"售后补发", "样品发货", "赠品", "退货损失"}
MAX_CONSUMER_ORDERS = 100
MAX_CONSUMER_REGIONS = 10

BASE_COLUMNS = [
    "订单编号", "原始子单号", "店铺名称", "订单来源", "订单状态", "订单类型",
    "订单退款状态", "订单明细退款状态", "交易时间", "付款时间", "发货时间", "结束时间",
    "客户唯一编码", "客户编号", "客户网名", "省", "市", "区", "省市县", "仓库",
    "商家编码", "货品编号", "货品名称", "规格名称", "分类", "赠品方式",
    "数量", "标价", "优惠", "优惠.1", "成交价", "分摊后价格", "分摊后总价", "分摊邮费",
    "订单邮费", "其它费用", "订单总优惠", "应收金额", "已付", "货到付款金额",
    "预估货品成本", "预估邮资成本", "货品当前成本", "成本价", "佣金",
]

NUMERIC_COLUMNS = {
    "数量", "标价", "优惠", "优惠.1", "成交价", "分摊后价格", "分摊后总价", "分摊邮费",
    "订单邮费", "其它费用", "订单总优惠", "应收金额", "已付", "货到付款金额",
    "预估货品成本", "预估邮资成本", "货品当前成本", "成本价", "佣金",
}


@dataclass
class FileResult:
    orders: pd.DataFrame
    products: pd.DataFrame
    quality: dict


def _first_valid(series: pd.Series):
    valid = series.dropna()
    if valid.empty:
        return np.nan
    text = valid.astype(str).str.strip()
    valid = valid[text.ne("") & text.ne("nan")]
    return valid.iloc[0] if not valid.empty else np.nan


def _safe_div(num: pd.Series, den: pd.Series) -> pd.Series:
    return num.div(den.replace(0, np.nan))


def _base_customer_identifier(df: pd.DataFrame, priority: list[str]) -> pd.Series:
    user = pd.Series(pd.NA, index=df.index, dtype="string")
    for customer_col in priority:
        if customer_col not in df:
            continue
        candidate = df[customer_col].astype("string").str.strip().replace({"": pd.NA, "nan": pd.NA})
        user = user.fillna(candidate)
    return user


def _county_address_from_orders(df: pd.DataFrame) -> pd.Series:
    """区县地址口径：优先原始 CU/CV/CW「省/市/区」，缺失或为“无”时回退 S「省市县」。"""
    def clean_geo(column: str) -> pd.Series:
        if column not in df:
            return pd.Series(pd.NA, index=df.index, dtype="string")
        return (
            df[column].astype("string").str.strip()
            .replace({"": pd.NA, "nan": pd.NA, "None": pd.NA, "无": pd.NA, "未知": pd.NA})
        )

    combined = clean_geo("省市县")
    tokens = combined.str.split()
    province = clean_geo("省").fillna(tokens.str[0])
    city = clean_geo("市").fillna(tokens.str[1])
    district = clean_geo("区").fillna(tokens.str[2])
    address = province.fillna("").str.cat(city.fillna(""), sep=" ").str.cat(district.fillna(""), sep=" ")
    address = address.str.replace(r"\s+", " ", regex=True).str.strip().replace({"": "未知"})
    return address


def _shared_customer_ids(df: pd.DataFrame, user: pd.Series) -> set[str]:
    """识别 To-C 场景下疑似平台共享/批量发货账号，避免把大量消费者合并成一个客户。"""
    identified = user.notna()
    if not identified.any():
        return set()
    check = pd.DataFrame({
        "客户标识": user.loc[identified],
        "订单编号": df.loc[identified, "订单编号"].astype("string"),
        "区县地址": _county_address_from_orders(df).loc[identified],
    })
    stats = check.groupby("客户标识", dropna=False).agg(
        订单数=("订单编号", "nunique"),
        地区数=("区县地址", "nunique"),
    )
    return set(stats.index[stats["订单数"].gt(MAX_CONSUMER_ORDERS) | stats["地区数"].gt(MAX_CONSUMER_REGIONS)].astype(str))


def _read_header(path: Path) -> list[str]:
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        cols = pd.read_excel(path, nrows=0).columns
    return [str(c).strip() for c in cols]


def discover_files(input_dir: Path, include_archive: bool) -> tuple[list[Path], list[Path]]:
    if not input_dir.exists():
        raise FileNotFoundError(f"输入目录不存在: {input_dir}")
    files = sorted(
        p for p in input_dir.glob("*.xlsx")
        if not p.name.startswith("~$") and not p.name.startswith(".")
    )
    active = [p for p in files if "归档" not in p.name]
    archived = [p for p in files if "归档" in p.name] if include_archive else []
    if not active and not archived:
        raise FileNotFoundError(f"未在 {input_dir} 找到 xlsx 文件")
    return active, archived


def _cache_path(cache_dir: Path, path: Path, sample_rows: int | None) -> Path:
    stat = path.stat()
    raw = f"{CACHE_VERSION}|{path.resolve()}|{stat.st_size}|{stat.st_mtime_ns}|{sample_rows}"
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return cache_dir / f"{path.stem}_{digest}.pkl"


def _line_discount(df: pd.DataFrame, archived: bool) -> pd.Series:
    # 归档表中 pandas 将订单优惠/明细优惠读为「优惠」/「优惠.1」。
    col = "优惠.1" if archived and "优惠.1" in df.columns else "优惠"
    return df[col] if col in df.columns else pd.Series(0.0, index=df.index)


def _order_discount(df: pd.DataFrame, archived: bool) -> pd.Series:
    col = "优惠" if archived and "订单总优惠" not in df.columns else "订单总优惠"
    return df[col] if col in df.columns else pd.Series(0.0, index=df.index)


def load_and_standardize(
    path: Path,
    source_kind: str,
    cache_dir: Path,
    sample_rows: int | None,
    use_cache: bool,
) -> FileResult:
    cache_dir.mkdir(parents=True, exist_ok=True)
    cp = _cache_path(cache_dir, path, sample_rows)
    if use_cache and cp.exists():
        payload = pd.read_pickle(cp)
        return FileResult(payload["orders"], payload["products"], payload["quality"])

    header = _read_header(path)
    usecols = [c for c in BASE_COLUMNS if c in header]
    required = {"订单编号", "店铺名称", "订单类型", "分摊后总价"}
    missing_required = sorted(required - set(usecols))
    if missing_required:
        raise ValueError(f"{path.name} 缺少必要字段: {missing_required}")

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        df = pd.read_excel(path, usecols=usecols, nrows=sample_rows)
    df.columns = [str(c).strip() for c in df.columns]
    raw_rows = len(df)

    for col in NUMERIC_COLUMNS.intersection(df.columns):
        df[col] = pd.to_numeric(df[col], errors="coerce")
    for col in PERIOD_COLUMNS.values():
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce")

    order_id = df["订单编号"].astype("string").str.strip()
    valid_order = order_id.notna() & order_id.ne("") & order_id.str.lower().ne("nan")
    blank_order_rows = int((~valid_order).sum())
    df = df.loc[valid_order].copy()
    df["订单编号"] = order_id.loc[valid_order]

    archived = source_kind == "已归档"
    df["明细优惠"] = pd.to_numeric(_line_discount(df, archived), errors="coerce").fillna(0)
    df["订单优惠_口径"] = pd.to_numeric(_order_discount(df, archived), errors="coerce").fillna(0)
    for col in ["数量", "标价", "分摊后总价", "分摊邮费", "货品当前成本", "成本价", "佣金"]:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    df["标价金额"] = df["标价"] * df["数量"]
    df["分析销售额"] = df["分摊后总价"] + df["分摊邮费"]
    current_cost = df["货品当前成本"].where(df["货品当前成本"].gt(0))
    standard_cost = (df["成本价"] * df["数量"]).where(df["成本价"].gt(0))
    df["行成本"] = current_cost.fillna(standard_cost)
    df["成本已覆盖销售额"] = df["分析销售额"].where(df["行成本"].notna(), 0)
    df["行成本"] = df["行成本"].fillna(0)

    fingerprint_cols = [
        c for c in ["订单编号", "原始子单号", "商家编码", "货品编号", "规格名称", "数量", "分摊后总价", "分摊邮费"]
        if c in df.columns
    ]
    before_dedup = len(df)
    df = df.drop_duplicates(subset=fingerprint_cols)
    duplicate_rows = before_dedup - len(df)

    first_cols = [
        "店铺名称", "订单来源", "订单状态", "订单类型", "订单退款状态", "订单明细退款状态",
        "交易时间", "付款时间", "发货时间", "结束时间", "客户唯一编码", "客户编号", "客户网名", "省", "市", "区", "省市县", "仓库",
    ]
    order_once_cols = ["订单邮费", "其它费用", "应收金额", "已付", "货到付款金额", "预估货品成本", "预估邮资成本", "订单优惠_口径"]
    additive_cols = ["数量", "标价金额", "分摊后总价", "分摊邮费", "分析销售额", "明细优惠", "行成本", "成本已覆盖销售额", "佣金"]

    agg: dict[str, object] = {c: _first_valid for c in first_cols if c in df.columns}
    agg.update({c: "max" for c in order_once_cols if c in df.columns})
    agg.update({c: "sum" for c in additive_cols})
    orders = df.groupby("订单编号", as_index=False, dropna=False).agg(agg)
    orders["数据来源"] = source_kind
    orders["源文件"] = path.name
    orders["明细行数"] = df.groupby("订单编号").size().reindex(orders["订单编号"]).to_numpy()

    # 行成本完全缺失时，才使用订单预估成本；避免与已有行成本双重计算。
    if "预估货品成本" not in orders:
        orders["预估货品成本"] = 0.0
    orders["成本金额"] = orders["行成本"].where(
        orders["成本已覆盖销售额"].abs().gt(0),
        pd.to_numeric(orders["预估货品成本"], errors="coerce").fillna(0),
    )
    fallback = orders["成本已覆盖销售额"].abs().eq(0) & orders["成本金额"].gt(0)
    orders.loc[fallback, "成本已覆盖销售额"] = orders.loc[fallback, "分析销售额"]

    product_keys = [c for c in ["订单编号", "商家编码", "货品编号", "货品名称", "规格名称", "分类", "赠品方式"] if c in df.columns]
    product_aggs = {c: "sum" for c in ["数量", "标价金额", "分摊后总价", "分摊邮费", "分析销售额", "明细优惠", "行成本", "成本已覆盖销售额", "佣金"]}
    products = df.groupby(product_keys, as_index=False, dropna=False).agg(product_aggs)
    products["数据来源"] = source_kind
    products["源文件"] = path.name

    tx = pd.to_datetime(df.get("交易时间"), errors="coerce") if "交易时间" in df else pd.Series(dtype="datetime64[ns]")
    quality = {
        "文件": path.name,
        "数据来源": source_kind,
        "原始读取行数": raw_rows,
        "空订单号行数": blank_order_rows,
        "有效明细行数": len(df),
        "文件内重复明细行": duplicate_rows,
        "订单数": orders["订单编号"].nunique(),
        "交易时间最早": tx.min(),
        "交易时间最晚": tx.max(),
        "成本覆盖销售额": float(orders["成本已覆盖销售额"].sum()),
        "分析销售额": float(orders["分析销售额"].sum()),
        "缺失的可选字段": "、".join(sorted(set(BASE_COLUMNS) - set(header))),
    }
    quality["成本覆盖率"] = (
        quality["成本覆盖销售额"] / quality["分析销售额"]
        if quality["分析销售额"] else np.nan
    )

    if use_cache:
        pd.to_pickle({"orders": orders, "products": products, "quality": quality}, cp)
    return FileResult(orders, products, quality)


def _periodize(orders: pd.DataFrame, basis: str) -> pd.DataFrame:
    out = orders.copy()
    period_col = PERIOD_COLUMNS[basis]
    if period_col not in out:
        out[period_col] = pd.NaT
    out["分析日期"] = pd.to_datetime(out[period_col], errors="coerce")
    out["分析日期来源"] = np.where(out["分析日期"].notna(), period_col, "")
    # 归档导出没有发货/结束时间。为避免订单被静默丢弃，按付款、交易、发货、结束的顺序回退，并保留日期来源供复核。
    for fallback_col in ["付款时间", "交易时间", "发货时间", "结束时间"]:
        if fallback_col == period_col or fallback_col not in out:
            continue
        fallback_values = pd.to_datetime(out[fallback_col], errors="coerce")
        mask = out["分析日期"].isna() & fallback_values.notna()
        out.loc[mask, "分析日期"] = fallback_values.loc[mask]
        out.loc[mask, "分析日期来源"] = f"{fallback_col}(回退)"
    out["分析月"] = out["分析日期"].dt.to_period("M").astype("string")
    out["客户标识"] = _base_customer_identifier(out, ["客户唯一编码", "客户编号", "客户网名"])
    shared_ids = _shared_customer_ids(out, out["客户标识"])
    out["疑似共享客户编号"] = out["客户标识"].astype("string").isin(shared_ids)
    out.loc[out["疑似共享客户编号"], "客户标识"] = pd.NA
    # 不再用订单号冒充客户：那会把每个缺失客户的订单都计为新客，系统性高估客户数。
    out["客户已识别"] = out["客户标识"].notna()
    out["是否商业订单"] = out["订单类型"].isin(COMMERCIAL_ORDER_TYPES)
    refund_order = out["订单退款状态"] if "订单退款状态" in out else pd.Series("", index=out.index)
    refund_line = out["订单明细退款状态"] if "订单明细退款状态" in out else pd.Series("", index=out.index)
    refund_text = refund_order.fillna("").astype(str) + "|" + refund_line.fillna("").astype(str)
    out["是否退款标记"] = refund_text.str.contains("退款", na=False)
    for col in ["应收金额", "已付", "订单邮费", "其它费用", "预估邮资成本", "佣金", "订单优惠_口径"]:
        if col not in out:
            out[col] = 0.0
        out[col] = pd.to_numeric(out[col], errors="coerce").fillna(0)
    out["应收差异"] = out["分析销售额"] - out["应收金额"]
    out["收款差额"] = out["应收金额"] - out["已付"]
    out["毛利额"] = out["分析销售额"] - out["成本金额"]
    out["贡献利润"] = out["毛利额"] - out["预估邮资成本"] - out["佣金"]
    out["退款标记销售额"] = out["分析销售额"].where(out["是否退款标记"], 0)
    out["已识别客户销售额"] = out["分析销售额"].where(out["客户已识别"], 0)
    return out


def _aggregate_finance(df: pd.DataFrame, keys: list[str]) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    g = df.groupby(keys, dropna=False)
    out = g.agg(
        订单数=("订单编号", "nunique"),
        客户数=("客户标识", "nunique"),
        未识别客户订单数=("客户已识别", lambda x: int((~x).sum())),
        已识别客户销售额=("已识别客户销售额", "sum"),
        商品数量=("数量", "sum"),
        标价金额=("标价金额", "sum"),
        商品收入=("分摊后总价", "sum"),
        邮费收入=("分摊邮费", "sum"),
        分析销售额=("分析销售额", "sum"),
        系统应收=("应收金额", "sum"),
        已付金额=("已付", "sum"),
        应收差异=("应收差异", "sum"),
        收款差额=("收款差额", "sum"),
        订单优惠=("订单优惠_口径", "sum"),
        成本金额=("成本金额", "sum"),
        成本已覆盖销售额=("成本已覆盖销售额", "sum"),
        预估邮资成本=("预估邮资成本", "sum"),
        佣金=("佣金", "sum"),
        毛利额=("毛利额", "sum"),
        贡献利润=("贡献利润", "sum"),
        退款标记订单数=("是否退款标记", "sum"),
        退款标记销售额=("退款标记销售额", "sum"),
    ).reset_index()
    out["客单价"] = _safe_div(out["分析销售额"], out["订单数"])
    out["人均销售额"] = _safe_div(out["已识别客户销售额"], out["客户数"])
    out["成本覆盖率"] = _safe_div(out["成本已覆盖销售额"], out["分析销售额"])
    out["毛利率"] = _safe_div(out["毛利额"], out["分析销售额"])
    out["贡献利润率"] = _safe_div(out["贡献利润"], out["分析销售额"])
    out["退款标记订单率"] = _safe_div(out["退款标记订单数"], out["订单数"])
    return out


LEGACY_SHEET_NAMES = [
    "A-0店铺销售额占比",
    "A-1_月均订单金额",
    "A-2_月均用户金额",
    "A-3_订单分层",
    "A-4_订单地址分布",
    "A-5_用户地址分布",
    "B-0_店铺详细对比",
    "B-1_用户价值分层",
    "B-2_用户地址统计",
    "B-3_用户购买频次",
]


def _platform_from_shop(shop: pd.Series) -> pd.Series:
    text = shop.fillna("未知店铺").astype(str).str.strip()
    known = ["抖音", "快手", "天猫", "京东", "拼多多", "微信", "鲸灵", "淘工厂"]
    result = pd.Series("其他", index=text.index, dtype="object")
    for platform in known:
        result = result.where(~text.str.contains(platform, na=False), platform)
    return result


def _legacy_user_identifier(df: pd.DataFrame) -> pd.Series:
    """兼容去年经营口径：客户网名→客户唯一编码→旺店通客户编号→订单号补缺。"""
    user = _base_customer_identifier(df, ["客户网名", "客户唯一编码", "客户编号"])
    shared_ids = _shared_customer_ids(df, user)
    user = user.mask(user.astype("string").isin(shared_ids))
    return user.fillna(df["订单编号"].astype("string"))


def _province_from_orders(df: pd.DataFrame) -> pd.Series:
    if "省" in df:
        province = df["省"].astype("string").str.strip().replace({"": pd.NA, "nan": pd.NA})
    else:
        province = pd.Series(pd.NA, index=df.index, dtype="string")
    if "省市县" in df:
        fallback = df["省市县"].astype("string").str.strip().str.split().str[0]
        province = province.fillna(fallback)
    return province.fillna("未知")


def build_legacy_outputs(orders: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """按去年交付工作簿的 10 张表名和字段口径生成兼容结果。"""
    df = orders.copy()
    df["交易时间"] = pd.to_datetime(df.get("交易时间"), errors="coerce")
    df = df.loc[df["交易时间"].notna()].copy()
    df["交易年月"] = df["交易时间"].dt.to_period("M").astype(str)
    df["用户标识_经营口径"] = _legacy_user_identifier(df)
    df["平台"] = _platform_from_shop(df["店铺名称"])
    df["省份"] = _province_from_orders(df)
    df["区县地址"] = _county_address_from_orders(df)

    # A-0：全期间店铺销售额占比。
    a0 = df.groupby("店铺名称", dropna=False).agg(
        平台订单数=("订单编号", "nunique"), 平台订单金额=("分析销售额", "sum")
    ).reset_index()
    a0["平台订单数占比"] = (_safe_div(a0["平台订单数"], pd.Series(a0["平台订单数"].sum(), index=a0.index)) * 100).round(2)
    a0["平台金额占比"] = (_safe_div(a0["平台订单金额"], pd.Series(a0["平台订单金额"].sum(), index=a0.index)) * 100).round(2)
    a0 = a0.sort_values("平台订单金额", ascending=False)

    # A-1：店铺×月的订单数、金额、客单价。
    a1 = df.groupby(["交易年月", "店铺名称"], dropna=False).agg(
        月订单数=("订单编号", "nunique"), 月订单金额=("分析销售额", "sum")
    ).reset_index()
    a1["月均订单金额"] = (_safe_div(a1["月订单金额"], a1["月订单数"])).round(2)
    a1 = a1.sort_values(["交易年月", "店铺名称"])

    # A-2：店铺×月的用户和订单金额。
    a2 = df.groupby(["交易年月", "店铺名称"], dropna=False).agg(
        月订单金额=("分析销售额", "sum"),
        月用户数=("用户标识_经营口径", "nunique"),
        最大订单金额=("分析销售额", "max"),
        最小订单金额=("分析销售额", "min"),
    ).reset_index()
    a2["月均用户金额"] = _safe_div(a2["月订单金额"], a2["月用户数"]).round(2)
    a2 = a2[["交易年月", "店铺名称", "月订单金额", "月用户数", "月均用户金额", "最大订单金额", "最小订单金额"]]

    # A-3：保留去年的固定金额分层。
    bins = [-np.inf, 50, 100, 200, 500, 1000, np.inf]
    labels = ["小额订单(<50元)", "中小额订单(50-100元)", "中额订单(100-200元)", "中大额订单(200-500元)", "大额订单(500-1000元)", "超大额订单(≥1000元)"]
    df["订单层级"] = pd.cut(df["分析销售额"], bins=bins, labels=labels, right=False)
    a3 = df.groupby(["交易年月", "平台", "店铺名称", "订单层级"], observed=True, dropna=False).agg(
        订单数=("订单编号", "nunique"), 总销售额=("分析销售额", "sum")
    ).reset_index()
    a3["平均订单金额"] = _safe_div(a3["总销售额"], a3["订单数"])
    a3["订单数占比"] = a3.groupby(["交易年月", "店铺名称"])["订单数"].transform(lambda x: x / x.sum() * 100)
    a3["销售额占比"] = a3.groupby(["交易年月", "店铺名称"])["总销售额"].transform(lambda x: x / x.sum() * 100 if x.sum() else 0)

    # A-4/A-5：全期间地域分布。
    a4 = df.groupby(["平台", "店铺名称", "省份"], dropna=False).agg(
        订单数=("订单编号", "nunique"), 销售额=("分析销售额", "sum")
    ).reset_index()
    a4["订单数占比"] = a4.groupby("店铺名称")["订单数"].transform(lambda x: x / x.sum() * 100)
    a4["销售额占比"] = a4.groupby("店铺名称")["销售额"].transform(lambda x: x / x.sum() * 100 if x.sum() else 0)
    a4 = a4.sort_values(["平台", "店铺名称", "销售额"], ascending=[True, True, False])

    a5 = df.groupby(["店铺名称", "平台", "区县地址"], dropna=False).agg(
        用户数=("用户标识_经营口径", "nunique")
    ).reset_index().rename(columns={"区县地址": "省市县"})
    a5["用户数占比"] = a5.groupby("店铺名称")["用户数"].transform(lambda x: x / x.sum() * 100)
    a5 = a5.sort_values(["平台", "店铺名称", "用户数"], ascending=[True, True, False])

    # B-0：店铺经营对比。
    b0 = df.groupby("店铺名称", dropna=False).agg(
        订单数=("订单编号", "nunique"), 订单总金额=("分析销售额", "sum"),
        平均订单金额=("分析销售额", "mean"), 最大订单金额=("分析销售额", "max"),
        最小订单金额=("分析销售额", "min"), 用户数=("用户标识_经营口径", "nunique"),
    ).reset_index()
    b0["人均订单金额"] = _safe_div(b0["订单总金额"], b0["用户数"])
    b0["人均订单数"] = _safe_div(b0["订单数"], b0["用户数"])
    b0 = b0.sort_values("订单总金额", ascending=False)

    # B-1/B-3：全期间用户价值和复购明细。
    user = df.groupby("用户标识_经营口径", dropna=False).agg(
        购买次数=("订单编号", "nunique"), 总金额=("分析销售额", "sum"),
        首次购买时间=("交易时间", "min"), 最后购买时间=("交易时间", "max"),
    ).reset_index().rename(columns={"用户标识_经营口径": "客户网名"})
    user["平均订单金额"] = _safe_div(user["总金额"], user["购买次数"])

    # 与去年口径完全一致，但改为向量化分类，避免全量数据逐用户 apply。
    freq = user["购买次数"]
    amount = user["总金额"]
    user["用户类型"] = np.select(
        [
            freq.ge(5) & amount.ge(1000),
            freq.ge(5),
            amount.ge(1000),
            freq.ge(3) & amount.ge(500),
            freq.ge(2),
        ],
        [
            "高价值用户(高频高额)",
            "高频用户(高频低额)",
            "高额用户(低频高额)",
            "中价值用户",
            "复购用户",
        ],
        default="新用户/单次购买",
    )
    b1 = user.groupby("用户类型").agg(
        用户数=("客户网名", "count"), 总购买次数=("购买次数", "sum"), 总金额=("总金额", "sum")
    ).reset_index()
    b1["平均购买次数"] = _safe_div(b1["总购买次数"], b1["用户数"])
    b1["平均消费金额"] = _safe_div(b1["总金额"], b1["用户数"])
    b1["用户数占比"] = b1["用户数"] / b1["用户数"].sum() * 100
    b1["金额占比"] = b1["总金额"] / b1["总金额"].sum() * 100 if b1["总金额"].sum() else 0
    b1 = b1.sort_values("总金额", ascending=False)

    # B-2/B-3 的“购买/复购”只统计正销售额订单，排除样品、赠品和零价发货造成的虚假频次。
    purchase_orders = df.loc[df["分析销售额"].gt(0)].copy()
    purchase_user = purchase_orders.groupby("用户标识_经营口径", dropna=False).agg(
        购买次数=("订单编号", "nunique"), 总金额=("分析销售额", "sum"),
        首次购买时间=("交易时间", "min"), 最后购买时间=("交易时间", "max"),
    ).reset_index().rename(columns={"用户标识_经营口径": "客户网名"})
    purchase_user["平均订单金额"] = _safe_div(purchase_user["总金额"], purchase_user["购买次数"])
    repeat_ids = set(purchase_user.loc[purchase_user["购买次数"] >= 2, "客户网名"].astype(str))
    repeat_orders = purchase_orders.loc[purchase_orders["用户标识_经营口径"].astype(str).isin(repeat_ids)]
    addr = repeat_orders.groupby("用户标识_经营口径").agg(
        收货地址数=("区县地址", "nunique"), 总订单数=("订单编号", "nunique")
    ).reset_index()
    b2 = addr.groupby("收货地址数").agg(用户数=("用户标识_经营口径", "count")).reset_index()
    b2["用户数占比"] = b2["用户数"] / b2["用户数"].sum() * 100
    b3 = purchase_user.loc[purchase_user["购买次数"] >= 2, ["客户网名", "购买次数", "总金额", "平均订单金额", "首次购买时间", "最后购买时间"]].sort_values(["购买次数", "总金额"], ascending=False)

    return dict(zip(LEGACY_SHEET_NAMES, [a0, a1, a2, a3, a4, a5, b0, b1, b2, b3]))


def _legacy_summary(sheet_name: str, df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    if sheet_name == "A-1_月均订单金额":
        return df.groupby("店铺名称").agg({"月订单数": "mean", "月订单金额": "mean", "月均订单金额": "mean"}).reset_index().rename(columns={"店铺名称": "行标签"})
    if sheet_name == "A-2_月均用户金额":
        return df.groupby("店铺名称").agg({"月订单金额": "mean", "月用户数": "mean", "月均用户金额": "mean"}).reset_index().rename(columns={"店铺名称": "行标签"})
    if sheet_name == "A-3_订单分层":
        out = df.groupby(["平台", "订单层级"], observed=True).agg(订单数=("订单数", "sum"), 总销售额=("总销售额", "sum")).reset_index()
        out["平均订单金额"] = _safe_div(out["总销售额"], out["订单数"])
        out["订单数占比"] = out.groupby("平台")["订单数"].transform(lambda x: x / x.sum() * 100)
        out["销售额占比"] = out.groupby("平台")["总销售额"].transform(lambda x: x / x.sum() * 100 if x.sum() else 0)
        return out
    if sheet_name == "A-4_订单地址分布":
        out = df.groupby(["平台", "省份"]).agg(订单数=("订单数", "sum"), 销售额=("销售额", "sum")).reset_index()
        out["订单数占比"] = out.groupby("平台")["订单数"].transform(lambda x: x / x.sum() * 100)
        out["销售额占比"] = out.groupby("平台")["销售额"].transform(lambda x: x / x.sum() * 100 if x.sum() else 0)
        return out
    if sheet_name == "A-5_用户地址分布":
        return df.groupby(["平台", "省市县"]).agg(用户数=("用户数", "sum")).reset_index()
    return pd.DataFrame()


def build_outputs(
    orders: pd.DataFrame,
    products: pd.DataFrame,
    basis: str,
    quality_rows: list[dict],
    excluded_1970: pd.DataFrame | None = None,
    period_filter_info: dict | None = None,
) -> dict[str, pd.DataFrame]:
    legacy_outputs = build_legacy_outputs(orders)
    orders = _periodize(orders, basis)
    valid_period = orders["分析月"].notna()
    commercial = orders.loc[valid_period & orders["是否商业订单"]].copy()

    monthly = _aggregate_finance(commercial, ["分析月"])
    shop_month = _aggregate_finance(commercial, ["分析月", "店铺名称"])
    order_type = _aggregate_finance(orders.loc[valid_period], ["分析月", "订单类型"])
    refund = _aggregate_finance(orders.loc[valid_period], ["分析月", "订单退款状态", "订单明细退款状态"])

    if products.empty or "订单编号" not in products:
        product = pd.DataFrame()
    else:
        product = products.merge(
            orders[["订单编号", "分析月", "店铺名称", "是否商业订单"]],
            on="订单编号", how="inner",
        )
        product = product.loc[product["是否商业订单"] & product["分析月"].notna()].copy()
    product_keys = [c for c in ["分析月", "商家编码", "货品编号", "货品名称", "规格名称", "分类"] if c in product.columns]
    if product.empty:
        product_summary = pd.DataFrame()
    else:
        product_summary = product.groupby(product_keys, dropna=False).agg(
            订单数=("订单编号", "nunique"), 数量=("数量", "sum"),
            分析销售额=("分析销售额", "sum"), 明细优惠=("明细优惠", "sum"),
            成本金额=("行成本", "sum"), 成本已覆盖销售额=("成本已覆盖销售额", "sum"),
        ).reset_index()
        product_summary["商品毛利额"] = product_summary["分析销售额"] - product_summary["成本金额"]
        product_summary["商品毛利率"] = _safe_div(product_summary["商品毛利额"], product_summary["分析销售额"])
        product_summary["成本覆盖率"] = _safe_div(product_summary["成本已覆盖销售额"], product_summary["分析销售额"])
        product_summary = product_summary.sort_values(["分析月", "分析销售额"], ascending=[True, False])

    q = pd.DataFrame(quality_rows)
    excluded_1970 = excluded_1970 if excluded_1970 is not None else pd.DataFrame()
    period_filter_info = period_filter_info or {}
    excluded_1970_sales = pd.to_numeric(excluded_1970.get("分析销售额", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
    cross_quality = pd.DataFrame([
        {"检查项": "订单级唯一性", "结果": "通过" if orders["订单编号"].is_unique else "需复核", "异常数": int(orders["订单编号"].duplicated().sum()), "说明": "财务事实表应每个订单仅一行"},
        {"检查项": "报告期间筛选", "结果": "通过", "异常数": int(period_filter_info.get("排除订单数", 0)), "说明": period_filter_info.get("说明", "未设置报告期间筛选")},
        {"检查项": "缺失分析期间", "结果": "通过" if orders["分析月"].notna().all() else "需复核", "异常数": int(orders["分析月"].isna().sum()), "说明": f"归期字段={PERIOD_COLUMNS[basis]}"},
        {"检查项": "分析日期使用回退字段", "结果": "信息披露", "异常数": int(orders["分析日期来源"].astype(str).str.contains("回退").sum()), "说明": "选定归期字段缺失时，按付款/交易/发货/结束时间回退"},
        {"检查项": "疑似共享客户编号订单", "结果": "信息披露", "异常数": int(orders["疑似共享客户编号"].sum()), "说明": f"客户编号>{MAX_CONSUMER_ORDERS}单或覆盖>{MAX_CONSUMER_REGIONS}个区县地址，按 To-C 口径不计入客户/复购"},
        {"检查项": "未识别客户订单", "结果": "通过" if commercial["客户已识别"].all() else ("需复核" if not commercial.empty and not commercial["客户已识别"].any() else "信息披露"), "异常数": int((~commercial["客户已识别"]).sum()), "说明": "缺少客户唯一编码/客户编号/客户网名的商业订单；不计入客户数、人均销售额"},
        {"检查项": "1970 异常日期", "结果": "通过" if not orders["分析日期"].dt.year.eq(1970).any() else "需复核", "异常数": int(orders["分析日期"].dt.year.eq(1970).sum()), "说明": "日期字段中发现系统默认时间的风险"},
        {"检查项": "1970 无效订单已剔除", "结果": "通过", "异常数": int(len(excluded_1970)), "说明": f"任一交易/付款/发货/结束时间年份为1970的订单整单剔除；剔除订单分析销售额={excluded_1970_sales:,.2f}元"},
        {"检查项": "应收对账差异", "结果": "信息披露", "异常数": int(orders["应收差异"].abs().gt(0.01).sum()), "说明": "分析销售额与系统应收差额绝对值>0.01的订单数"},
        {"检查项": "成本覆盖率", "结果": "信息披露", "异常数": np.nan, "说明": f"{(_safe_div(pd.Series([commercial['成本已覆盖销售额'].sum()]), pd.Series([commercial['分析销售额'].sum()])).iloc[0] if not commercial.empty else np.nan):.2%}"},
    ])

    definitions = pd.DataFrame([
        ["报告期间", period_filter_info.get("报告期间", "未限定"), "以交易时间限定整体订单范围；期间外或交易时间缺失的订单不进入分析"],
        ["分析期间", PERIOD_COLUMNS[basis], "可用 --period-basis 切换；缺失时按付款/交易/发货/结束时间回退，并在订单事实表标注来源"],
        ["商业订单", "订单类型∈网店销售/分销订单/线下订单", "补发、样品、赠品、退货损失不进入主 KPI"],
        ["客户数/人均销售额", "客户唯一编码 → 旺店通客户编号 → 客户网名", "缺失客户不用订单号冒充；人均销售额分子也只包含已识别客户销售额"],
        ["区县地址", "CU省 + CV市 + CW区", "CW区为空、无或未知时，才回退到 S省市县的对应字段；T地址为脱敏星号，不参与统计"],
        ["1970 无效日期", "交易/付款/发货/结束时间任一字段年份=1970", "视为系统默认或错误日期，整单不进入经营、客户及财务分析；剔除记录见Q-2"],
        ["疑似共享客户编号", f"客户编号>{MAX_CONSUMER_ORDERS}单或覆盖>{MAX_CONSUMER_REGIONS}个区县地址", "To-C 客户分析中排除，避免平台共享/批量发货账号造成虚假复购"],
        ["分析销售额", "∑分摊后总价 + ∑分摊邮费", "明细行口径，与系统应收并列对账"],
        ["系统应收/已付/订单优惠", "每订单仅取一次", "禁止在商品行上求和"],
        ["成本金额", "货品当前成本 → 成本价×数量 → 订单预估货品成本", "必须与成本覆盖率一起使用"],
        ["毛利额", "分析销售额-成本金额", "成本未完整时只是不完整估算"],
        ["贡献利润", "毛利额-预估邮资成本-佣金", "当前抽样中邮资成本/佣金大量为 0，不可视为完整贡献利润"],
        ["退款标记销售额", "带退款状态订单的分析销售额", "不是实际退款额，不冲减销售额"],
    ], columns=["指标", "口径", "财务解读/限制"])

    supplemental_outputs = {
        "F-0_月度财务总览": monthly,
        "F-1_店铺月度财务": shop_month,
        "F-2_商品毛利分析": product_summary,
        "F-3_订单类型分析": order_type,
        "F-4_退款风险披露": refund,
        "Q-0_数据质量总览": cross_quality,
        "Q-1_文件质量明细": q,
        "Q-2_1970剔除明细": excluded_1970,
        "D-0_财务口径字典": definitions,
    }
    return {**legacy_outputs, **supplemental_outputs}, orders


def _style_workbook(path: Path, report_date: pd.Timestamp | None):
    import unicodedata
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = load_workbook(path)
    finance_header_fill = PatternFill("solid", fgColor="1F4E78")
    legacy_header_fill = PatternFill("solid", fgColor="0070C0")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    platform_fills = {
        "抖音": PatternFill("solid", fgColor="E7E6E6"),
        "快手": PatternFill("solid", fgColor="FCE4D6"),
        "天猫": PatternFill("solid", fgColor="DDEBF7"),
        "京东": PatternFill("solid", fgColor="F4CCCC"),
        "拼多多": PatternFill("solid", fgColor="FFF2CC"),
        "微信": PatternFill("solid", fgColor="E2F0D9"),
    }
    thin = Side(style="thin", color="000000")
    report_date = pd.Timestamp(report_date).normalize() if report_date is not None and pd.notna(report_date) else pd.Timestamp.today().normalize()
    prepared_date = pd.Timestamp.today().normalize()

    def display_width(value) -> int:
        text = str(value) if value is not None else ""
        return sum(2 if unicodedata.east_asian_width(ch) in {"W", "F", "A"} else 1 for ch in text)

    for ws in wb.worksheets:
        is_legacy = ws.title in LEGACY_SHEET_NAMES
        header_row = 8 if ws.title.startswith("A-") and ws.title != "A-0店铺销售额占比" else 7 if is_legacy else 1
        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

        if is_legacy:
            data_headers = [str(ws.cell(header_row, c).value or "") for c in range(1, ws.max_column + 1)]
            blank_gap = next((c for c, v in enumerate(data_headers, 1) if not v), ws.max_column + 1)
            primary_cols = blank_gap - 1
            meta_col = max(primary_cols + 3, 8)
            ws["A1"] = "妙可蓝多食品科技股份有限公司"
            ws["A2"] = f"资产负债表日Balance sheet date：{report_date.year}年{report_date.month}月{report_date.day}日"
            ws["A3"] = "以人民币表示 Expressed in RMB"
            ws["A4"] = f"趋势分析-{ws.title}"
            if ws.title in {"B-2_用户地址统计", "B-3_用户购买频次"}:
                customer_note = f"客户口径：客户网名 → 客户唯一编码 → 旺店通客户编号；>{MAX_CONSUMER_ORDERS}单或>{MAX_CONSUMER_REGIONS}个区县地址的共享账号不计复购"
                address_note = "；地址口径：CU省 + CV市 + CW区（CW缺失/为无时回退S省市县）" if ws.title == "B-2_用户地址统计" else ""
                ws["A5"] = customer_note + address_note
                if primary_cols > 1:
                    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=primary_cols)
                ws["A5"].font = Font(name="微软雅黑", size=9, italic=True, color="7F6000")
            ws.cell(1, meta_col, "索引号：")
            ws.cell(1, meta_col + 1, "A-1")
            ws.cell(2, meta_col, "编制人：")
            ws.cell(2, meta_col + 1, "LL")
            ws.cell(3, meta_col, "编制日期：")
            ws.cell(3, meta_col + 1, prepared_date.to_pydatetime())
            ws.cell(4, meta_col, "复核日期：")
            ws.cell(4, meta_col + 1, prepared_date.to_pydatetime() + pd.Timedelta(days=2))
            ws["A6"] = "1. 数据分布"
            if header_row == 8:
                ws["A7"] = "分月陈列"
                if primary_cols > 1:
                    ws.merge_cells(start_row=7, start_column=1, end_row=7, end_column=primary_cols)
                summary_start = primary_cols + 3
                if summary_start <= ws.max_column and ws.cell(8, summary_start).value:
                    ws.cell(7, summary_start, "汇总数据")
                    ws.merge_cells(start_row=7, start_column=summary_start, end_row=7, end_column=ws.max_column)
            for row in range(1, 5):
                ws.cell(row, 1).font = Font(name="微软雅黑", size=10, bold=True)
                ws.cell(row, meta_col).font = Font(name="微软雅黑", size=10, bold=True)
                ws.cell(row, meta_col + 1).font = Font(name="微软雅黑", size=10, bold=True)
            ws["A6"].font = Font(name="微软雅黑", size=10, bold=True)
            if header_row == 8:
                for c in range(1, ws.max_column + 1):
                    ws.cell(7, c).fill = section_fill
                    ws.cell(7, c).font = Font(name="微软雅黑", size=10, bold=True)

        for cell in ws[header_row]:
            if cell.value is None:
                continue
            cell.fill = legacy_header_fill if is_legacy else finance_header_fill
            cell.font = Font(name="微软雅黑", color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

        scan_rows = min(ws.max_row, 5000 if is_legacy else 3000)
        for col_idx, col_cells in enumerate(ws.iter_cols(min_row=header_row, max_row=scan_rows), 1):
            header = str(ws.cell(header_row, col_idx).value or "")
            width = min(40, max(10, max((display_width(c.value) for c in col_cells), default=0) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
            if "率" in header or "占比" in header:
                fmt = "0.00" if is_legacy else "0.00%"
                for c in ws.iter_rows(min_row=header_row + 1, min_col=col_idx, max_col=col_idx):
                    c[0].number_format = fmt
            elif any(k in header for k in ["金额", "收入", "成本", "利润", "客单价", "人均"]):
                for c in ws.iter_rows(min_row=header_row + 1, min_col=col_idx, max_col=col_idx):
                    c[0].number_format = "#,##0.00"

        if is_legacy:
            ws.column_dimensions[get_column_letter(meta_col)].width = max(ws.column_dimensions[get_column_letter(meta_col)].width or 0, 12)
            ws.column_dimensions[get_column_letter(meta_col + 1)].width = max(ws.column_dimensions[get_column_letter(meta_col + 1)].width or 0, 16)
            header_map = {str(ws.cell(header_row, c).value): c for c in range(1, ws.max_column + 1)}
            platform_col = header_map.get("平台")
            shop_col = header_map.get("店铺名称")
            primary_end = next((c - 1 for c in range(1, ws.max_column + 1) if ws.cell(header_row, c).value is None), ws.max_column)
            for row in range(header_row + 1, ws.max_row + 1):
                platform = str(ws.cell(row, platform_col).value or "") if platform_col else ""
                if not platform and shop_col:
                    platform = str(ws.cell(row, shop_col).value or "")
                fill = next((v for k, v in platform_fills.items() if k in platform), None)
                for col in range(1, primary_end + 1):
                    cell = ws.cell(row, col)
                    cell.font = Font(name="微软雅黑", size=10)
                    cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
                    if fill:
                        cell.fill = fill
            ws.sheet_view.showGridLines = True
        else:
            ws.sheet_view.showGridLines = False
    wb.save(path)


def write_excel(outputs: dict[str, pd.DataFrame], output_path: Path, orders: pd.DataFrame, include_order_detail: bool):
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for name, df in outputs.items():
            if isinstance(df, pd.DataFrame):
                if name in LEGACY_SHEET_NAMES:
                    startrow = 7 if name.startswith("A-") and name != "A-0店铺销售额占比" else 6
                    df.to_excel(writer, sheet_name=name[:31], index=False, startrow=startrow)
                    summary = _legacy_summary(name, df)
                    if not summary.empty:
                        summary.to_excel(writer, sheet_name=name[:31], index=False, startrow=7, startcol=len(df.columns) + 2)
                else:
                    df.to_excel(writer, sheet_name=name[:31], index=False)
        if include_order_detail:
            if len(orders) > 1_048_575:
                raise ValueError("订单明细超过 Excel 单表行数上限，请去掉 --include-order-detail")
            orders.to_excel(writer, sheet_name="B-0_订单财务事实", index=False)
    report_date = pd.to_datetime(orders.get("交易时间"), errors="coerce").max() if "交易时间" in orders else None
    _style_workbook(output_path, report_date)


def run(args) -> Path:
    active_files, archive_files = discover_files(args.input, args.include_archive)
    print(f"发现主清单 {len(active_files)} 个，归档清单 {len(archive_files)} 个")
    print(f"归期口径: {PERIOD_COLUMNS[args.period_basis]}")
    if args.sample_rows:
        print(f"样本模式: 每个文件最多读取 {args.sample_rows:,} 行，输出不可作为全量财务结果")

    order_parts: list[pd.DataFrame] = []
    product_parts: list[pd.DataFrame] = []
    quality: list[dict] = []
    active_order_ids: set[str] = set()

    for i, path in enumerate(active_files, 1):
        print(f"[{i}/{len(active_files)}] 读取主清单: {path.name}")
        r = load_and_standardize(path, "主清单", args.cache, args.sample_rows, not args.no_cache)
        order_parts.append(r.orders)
        product_parts.append(r.products)
        quality.append(r.quality)
        active_order_ids.update(r.orders["订单编号"].astype(str))

    archive_overlap = 0
    for i, path in enumerate(archive_files, 1):
        print(f"[{i}/{len(archive_files)}] 读取归档清单: {path.name}")
        r = load_and_standardize(path, "已归档", args.cache, args.sample_rows, not args.no_cache)
        keep_orders = ~r.orders["订单编号"].astype(str).isin(active_order_ids)
        keep_ids = set(r.orders.loc[keep_orders, "订单编号"].astype(str))
        archive_overlap += int((~keep_orders).sum())
        order_parts.append(r.orders.loc[keep_orders].copy())
        product_parts.append(r.products.loc[r.products["订单编号"].astype(str).isin(keep_ids)].copy())
        r.quality["与主清单重叠订单数"] = int((~keep_orders).sum())
        quality.append(r.quality)

    nonempty_orders = [part for part in order_parts if not part.empty]
    nonempty_products = [part for part in product_parts if not part.empty]
    if not nonempty_orders:
        raise ValueError("所有文件读取后都没有有效订单")
    orders = pd.concat(nonempty_orders, ignore_index=True)
    products = pd.concat(nonempty_products, ignore_index=True) if nonempty_products else pd.DataFrame()

    # 主清单分卷之间如有订单边界跨文件，在这里再聚合。
    duplicate_mask = orders["订单编号"].duplicated(keep=False)
    duplicate_order_parts = int(duplicate_mask.sum())
    if duplicate_order_parts:
        print(f"提示: {duplicate_order_parts} 条文件级订单记录存在跨分卷重复/边界，已按订单再次归并")
        first_cols = [c for c in orders.columns if c not in {"订单编号", "数量", "标价金额", "分摊后总价", "分摊邮费", "分析销售额", "明细优惠", "行成本", "成本已覆盖销售额", "成本金额", "佣金", "明细行数"}]
        additive = [c for c in ["数量", "标价金额", "分摊后总价", "分摊邮费", "分析销售额", "明细优惠", "行成本", "成本已覆盖销售额", "成本金额", "佣金", "明细行数"] if c in orders]
        agg = {c: _first_valid for c in first_cols}
        # 成本金额稍后根据归并后的行成本/订单预估成本重新计算，不直接相加文件级回退值。
        agg.update({c: "sum" for c in additive if c != "成本金额"})
        # 仅对真正跨分卷重复的极少量订单执行自定义聚合；其余订单直接保留。
        # 避免对数百万唯一订单逐组调用 Python 函数，结果与全表 groupby 等价。
        unique_orders = orders.loc[~duplicate_mask]
        merged_duplicates = (
            orders.loc[duplicate_mask]
            .groupby("订单编号", as_index=False, sort=False)
            .agg(agg)
        )
        orders = pd.concat([unique_orders, merged_duplicates], ignore_index=True)
        if "预估货品成本" not in orders:
            orders["预估货品成本"] = 0.0
        orders["成本金额"] = orders["行成本"].where(
            orders["成本已覆盖销售额"].abs().gt(0),
            pd.to_numeric(orders["预估货品成本"], errors="coerce").fillna(0),
        )

    quality.append({
        "文件": "跨文件总检查", "数据来源": "合并后", "原始读取行数": np.nan,
        "空订单号行数": np.nan, "有效明细行数": int(sum(q.get("有效明细行数", 0) for q in quality)),
        "文件内重复明细行": int(sum(q.get("文件内重复明细行", 0) for q in quality)),
        "订单数": orders["订单编号"].nunique(), "与主清单重叠订单数": archive_overlap,
        "跨分卷重复/边界记录": duplicate_order_parts,
    })

    excluded_1970 = pd.DataFrame()
    if args.exclude_1970:
        date_columns = [c for c in ["交易时间", "付款时间", "发货时间", "结束时间"] if c in orders]
        invalid_by_field = {
            c: pd.to_datetime(orders[c], errors="coerce").dt.year.eq(1970)
            for c in date_columns
        }
        invalid_1970 = pd.concat(invalid_by_field, axis=1).any(axis=1) if invalid_by_field else pd.Series(False, index=orders.index)
        if invalid_1970.any():
            excluded_1970 = orders.loc[invalid_1970].copy()
            excluded_1970["1970异常字段"] = [
                "、".join(c for c, mask in invalid_by_field.items() if bool(mask.loc[idx]))
                for idx in excluded_1970.index
            ]
            detail_columns = [
                "订单编号", "数据来源", "源文件", "店铺名称", "订单类型", "分析销售额",
                "1970异常字段", "交易时间", "付款时间", "发货时间", "结束时间",
            ]
            excluded_1970 = excluded_1970[[c for c in detail_columns if c in excluded_1970]].reset_index(drop=True)
            excluded_ids = set(orders.loc[invalid_1970, "订单编号"].astype(str))
            orders = orders.loc[~invalid_1970].copy()
            if not products.empty:
                products = products.loc[~products["订单编号"].astype(str).isin(excluded_ids)].copy()
            print(f"已剔除 1970 无效日期订单: {len(excluded_ids)} 单")

    period_filter_info = {}
    if args.start_date or args.end_date:
        if "交易时间" not in orders:
            raise ValueError("设置报告期间时必须存在交易时间字段")
        start_date = pd.Timestamp(args.start_date).normalize() if args.start_date else pd.Timestamp.min.normalize()
        end_date = pd.Timestamp(args.end_date).normalize() if args.end_date else pd.Timestamp.max.normalize()
        if start_date > end_date:
            raise ValueError("报告期间开始日期不能晚于结束日期")
        trade_date = pd.to_datetime(orders["交易时间"], errors="coerce")
        in_period = trade_date.ge(start_date) & trade_date.lt(end_date + pd.Timedelta(days=1))
        excluded_period_orders = orders.loc[~in_period]
        kept_order_ids = set(orders.loc[in_period, "订单编号"].astype(str))
        period_filter_info = {
            "报告期间": f"{start_date:%Y-%m-%d} 至 {end_date:%Y-%m-%d}",
            "排除订单数": int((~in_period).sum()),
            "说明": (
                f"以交易时间限定 {start_date:%Y-%m-%d} 至 {end_date:%Y-%m-%d}；"
                f"期间外或交易时间缺失共排除 {int((~in_period).sum()):,} 单，"
                f"其分析销售额合计 {pd.to_numeric(excluded_period_orders['分析销售额'], errors='coerce').fillna(0).sum():,.2f} 元"
            ),
        }
        orders = orders.loc[in_period].copy()
        if not products.empty:
            products = products.loc[products["订单编号"].astype(str).isin(kept_order_ids)].copy()
        print(f"报告期间 {start_date:%Y-%m-%d} 至 {end_date:%Y-%m-%d}：保留 {len(orders):,} 单，排除 {int((~in_period).sum()):,} 单")

    outputs, periodized_orders = build_outputs(
        orders, products, args.period_basis, quality, excluded_1970, period_filter_info
    )
    suffix = "_样本" if args.sample_rows else ""
    suffix += "_剔除1970" if args.exclude_1970 else ""
    if args.start_date or args.end_date:
        start_tag = pd.Timestamp(args.start_date).strftime("%Y%m%d") if args.start_date else "开始"
        end_tag = pd.Timestamp(args.end_date).strftime("%Y%m%d") if args.end_date else "结束"
        suffix += f"_{start_tag}-{end_tag}"
    output_path = args.output / f"Bespoke DA routine_平台经营数据统计_补充财务{suffix}.xlsx"
    write_excel(outputs, output_path, periodized_orders, args.include_order_detail)
    print(f"完成: {output_path}")
    return output_path


def parse_args(argv: Iterable[str] | None = None):
    parser = argparse.ArgumentParser(description="旺店通订单财务分析（平铺分卷版）")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="订单 xlsx 所在的平铺目录")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="输出目录")
    parser.add_argument("--cache", type=Path, default=DEFAULT_CACHE, help="文件级缓存目录")
    parser.add_argument("--period-basis", choices=sorted(PERIOD_COLUMNS), default="shipment", help="归期口径")
    parser.add_argument("--sample-rows", type=int, default=None, help="每个文件仅读取前 N 行，只用于测试")
    parser.add_argument("--exclude-archive", dest="include_archive", action="store_false", help="不纳入已归档订单")
    parser.set_defaults(include_archive=True)
    parser.add_argument("--include-order-detail", action="store_true", help="在结果中附订单级事实表（可能较大）")
    parser.add_argument("--exclude-1970", action="store_true", help="剔除任一主要日期字段年份为 1970 的整张订单")
    parser.add_argument("--start-date", type=str, default=None, help="报告期间开始日期（YYYY-MM-DD，按交易时间筛选）")
    parser.add_argument("--end-date", type=str, default=None, help="报告期间结束日期（YYYY-MM-DD，按交易时间筛选）")
    parser.add_argument("--no-cache", action="store_true", help="不读写文件级 PKL 缓存")
    return parser.parse_args(argv)


if __name__ == "__main__":
    try:
        run(parse_args())
    except Exception as exc:
        print(f"运行失败: {type(exc).__name__}: {exc}", file=sys.stderr)
        traceback.print_exc()
        raise SystemExit(1)
